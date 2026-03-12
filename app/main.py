from __future__ import annotations

import csv
import json
import os
import platform
import subprocess
import sys
import threading
import tempfile
import time
from datetime import datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from urllib.parse import urlparse
from urllib.request import urlopen

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from .barcode_utils import generate_barcode
from .config import APP_PATCH, APP_TITLE, APP_VERSION, BACKUPS_DIR, EXPORTS_DIR
from .db import fetch_all
from .services import StoreService
from .ui_assets import ICON_GLYPHS, LEVEL_COLORS, resolve_icon_paths


THEME_PRESETS = {
    "Clair Pro": {
        "bg": "#eef2f7",
        "surface": "#ffffff",
        "header": "#0d5db8",
        "sidebar": "#0a4a95",
        "primary": "#156fd1",
        "success": "#1f8f4a",
        "danger": "#c73f35",
        "warning": "#e09a1a",
        "text_light": "#ffffff",
        "text_dark": "#1f2a35",
    },
    "Nuit Operateur": {
        "bg": "#151c2b",
        "surface": "#1f2738",
        "header": "#0b111f",
        "sidebar": "#0e1628",
        "primary": "#2f85ff",
        "success": "#31a95b",
        "danger": "#de5a50",
        "warning": "#f0b33a",
        "text_light": "#f5f7ff",
        "text_dark": "#dce5f7",
    },
    "Contraste Eleve": {
        "bg": "#ffffff",
        "surface": "#ffffff",
        "header": "#000000",
        "sidebar": "#111111",
        "primary": "#0057ff",
        "success": "#008a38",
        "danger": "#b00020",
        "warning": "#b26b00",
        "text_light": "#ffffff",
        "text_dark": "#111111",
    },
}

PALETTE = THEME_PRESETS["Clair Pro"].copy()

CURRENCY_CHOICES: list[tuple[str, str]] = [
    ("EUR", "Euro"),
    ("USD", "Dollar americain"),
    ("CDF", "Franc congolais (RDC)"),
    ("XAF", "Franc CFA BEAC (Afrique centrale)"),
    ("XOF", "Franc CFA BCEAO (Afrique de l'Ouest)"),
    ("MAD", "Dirham marocain"),
    ("DZD", "Dinar algerien"),
    ("TND", "Dinar tunisien"),
    ("EGP", "Livre egyptienne"),
    ("GHS", "Cedi ghaneen"),
    ("NGN", "Naira nigerian"),
    ("KES", "Shilling kenyan"),
    ("UGX", "Shilling ougandais"),
    ("TZS", "Shilling tanzanien"),
    ("ZAR", "Rand sud-africain"),
    ("RWF", "Franc rwandais"),
    ("BIF", "Franc burundais"),
    ("ZMW", "Kwacha zambien"),
    ("MZN", "Metical mozambicain"),
]

STARTUP_TIPS = [
    "Astuce: utilisez la recherche dans Produits pour aller vite.",
    "Astuce: scannez le code-barres puis Entrer pour ajouter au panier.",
    "Astuce: activez la sauvegarde avant une restauration complete.",
    "Astuce: changez la devise depuis Parametres > Design et branding.",
    "Astuce: exportez vos rapports en PDF et Excel pour l'audit.",
]


class StartupSplash(tk.Toplevel):
    def __init__(self, parent: tk.Tk, app_title: str):
        super().__init__(parent)
        self.title("Chargement")
        self.geometry("760x470")
        self.resizable(False, False)
        self.overrideredirect(True)
        self.configure(bg="#0b4db3")
        self.attributes("-topmost", True)

        parent.update_idletasks()
        x = max(20, (parent.winfo_screenwidth() - 760) // 2)
        y = max(20, (parent.winfo_screenheight() - 470) // 2)
        self.geometry(f"760x470+{x}+{y}")

        self.canvas = tk.Canvas(self, width=760, height=470, highlightthickness=0, bg="#0b4db3")
        self.canvas.pack(fill="both", expand=True)
        self._draw_background()
        self.lift()
        self.update_idletasks()

        self.canvas.create_text(380, 54, text=app_title, fill="#ffffff", font=("Segoe UI", 18, "bold"))
        self.canvas.create_text(380, 88, text="Initialisation en cours", fill="#dce8ff", font=("Segoe UI", 12))

        self.status_var = tk.StringVar(value="Preparation...")
        self.tip_var = tk.StringVar(value=STARTUP_TIPS[0])
        self.percent_var = tk.StringVar(value="0%")

        self.status_label = tk.Label(self, textvariable=self.status_var, bg="#0f5ad0", fg="#eaf2ff", font=("Segoe UI", 13, "bold"))
        self.tip_label = tk.Label(self, textvariable=self.tip_var, bg="#0f5ad0", fg="#b7ccf4", font=("Segoe UI", 10), wraplength=640, justify="center")
        self.percent_label = tk.Label(self, textvariable=self.percent_var, bg="#0f5ad0", fg="#ffffff", font=("Segoe UI", 20, "bold"))

        self.canvas.create_window(380, 350, window=self.status_label)
        self.canvas.create_window(380, 382, window=self.tip_label)

        self.canvas.create_rectangle(90, 404, 670, 432, fill="#153f92", outline="#153f92", width=0)
        self.progress_fill = self.canvas.create_rectangle(92, 406, 92, 430, fill="#ffd644", outline="#ffd644", width=0)
        self.canvas.create_window(380, 442, window=self.percent_label)

    def _draw_background(self):
        # Layered blue background inspired by the provided mockup.
        for i, color in enumerate(["#0b4db3", "#0e56c6", "#0f5ad0", "#115fd9"]):
            self.canvas.create_rectangle(0, i * 120, 760, (i + 1) * 120, fill=color, outline=color)
        self.canvas.create_oval(-120, 220, 180, 500, fill="#164aa5", outline="")
        self.canvas.create_oval(540, 260, 860, 560, fill="#1354b7", outline="")
        self.canvas.create_oval(250, 140, 510, 390, fill="#0f57c8", outline="")
        self.canvas.create_text(380, 222, text="Progression du demarrage", fill="#dbe8ff", font=("Segoe UI", 14, "bold"))

    def update_progress(self, percent: int, status: str, tip: str):
        p = max(0, min(100, int(percent)))
        x2 = 92 + int((576 * p) / 100)
        self.canvas.coords(self.progress_fill, 92, 406, x2, 430)
        self.status_var.set(status)
        self.tip_var.set(tip)
        self.percent_var.set(f"{p}%")
        self.update_idletasks()

class StoreApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.withdraw()
        self.title(APP_TITLE)
        self.geometry("1360x820")
        self.minsize(980, 620)
        self.configure(bg=PALETTE["bg"])
        self.setup_styles()
        self.service = StoreService()
        self.startup_health = {"ok": True, "issues": [], "counts": {}}
        self.update_info = {"enabled": False, "available": False}
        self.current_user = None

        self._container = ttk.Frame(self)
        self._container.pack(fill="both", expand=True)

        self._run_startup_sequence()
        self.show_login()
        self.deiconify()
        self.lift()

    def _run_startup_sequence(self):
        splash = StartupSplash(self, APP_TITLE)
        self.update_idletasks()

        steps = [
            ("Connexion au service applicatif", lambda: None),
            ("Verification des extensions de base", self.service.ensure_extensions),
            ("Verification du compte administrateur", self.service.ensure_default_admin),
            ("Controle integrite des donnees", self._load_startup_health),
            ("Recherche des mises a jour", self._load_update_info),
            ("Creation sauvegarde automatique", self.service.create_auto_backup_if_needed),
            ("Finalisation de l'interface", lambda: None),
        ]

        total = len(steps)
        current_pct = 0.0
        try:
            for idx, (label, action) in enumerate(steps, start=1):
                target_pct = float((idx / total) * 100)
                tip = STARTUP_TIPS[(idx - 1) % len(STARTUP_TIPS)]

                # Animate while the step is executing so slow machines still feel alive.
                step_start = time.perf_counter()
                current_pct = self._run_step_with_pulse(splash, action, current_pct, target_pct, label, tip)
                step_elapsed = time.perf_counter() - step_start

                # Fast machines get a slightly longer easing, slow machines get shorter easing.
                ease_duration = 0.30 if step_elapsed < 0.25 else (0.22 if step_elapsed < 1.2 else 0.14)
                current_pct = self._animate_splash_progress(splash, current_pct, target_pct, label, tip, ease_duration)

            self._animate_splash_progress(splash, current_pct, 100.0, "Demarrage termine", STARTUP_TIPS[-1], 0.26)
        finally:
            splash.attributes("-topmost", False)
            splash.destroy()

    def _animate_splash_progress(self, splash: StartupSplash, start_pct: float, end_pct: float, status: str, tip: str, duration: float) -> float:
        frames = max(1, int(duration / 0.03))
        for i in range(1, frames + 1):
            pct = start_pct + ((end_pct - start_pct) * (i / frames))
            splash.update_progress(int(pct), status, tip)
            self.update()
            time.sleep(0.03)
        return float(end_pct)

    def _run_step_with_pulse(self, splash: StartupSplash, action, start_pct: float, end_pct: float, status: str, tip: str) -> float:
        outcome = {"error": None}

        def worker():
            try:
                action()
            except Exception as exc:
                outcome["error"] = exc

        thread = threading.Thread(target=worker, daemon=True)
        thread.start()

        pulse_min = float(start_pct)
        pulse_max = max(pulse_min + 1.0, float(end_pct) - 2.0)
        pulse = pulse_min
        direction = 1.0

        while thread.is_alive():
            splash.update_progress(int(pulse), status, tip)
            self.update()
            time.sleep(0.03)

            pulse += direction
            if pulse >= pulse_max:
                pulse = pulse_max
                direction = -1.0
            elif pulse <= pulse_min:
                pulse = pulse_min
                direction = 1.0

        thread.join()
        if outcome["error"] is not None:
            raise outcome["error"]
        return float(max(start_pct, min(pulse, end_pct - 1.0)))

    def _load_startup_health(self):
        self.startup_health = self.service.startup_integrity_check()

    def _load_update_info(self):
        self.update_info = self.service.check_remote_update()

    def apply_theme(self, theme_name: str):
        preset = THEME_PRESETS.get(theme_name)
        if not preset:
            return
        PALETTE.clear()
        PALETTE.update(preset)
        self.configure(bg=PALETTE["bg"])
        self.setup_styles()

    def clear_container(self):
        for child in self._container.winfo_children():
            child.destroy()

    def show_login(self):
        self.clear_container()
        LoginView(self._container, self).pack(fill="both", expand=True)
        if not self.startup_health.get("ok", True):
            issues = "\n".join(f"- {x}" for x in self.startup_health.get("issues", [])[:6])
            messagebox.showwarning("Integrite des donnees", f"Des anomalies ont ete detectees au demarrage:\n{issues}")

        if self.update_info.get("enabled") and self.update_info.get("available"):
            latest = self.update_info.get("latest", "?")
            latest_patch = int(self.update_info.get("latest_patch", 0) or 0)
            notes = self.update_info.get("notes", "")
            current_label = f"{APP_VERSION} (patch {APP_PATCH})"
            latest_label = f"{latest} (patch {latest_patch})" if latest_patch > 0 else str(latest)
            do_update = messagebox.askyesno(
                "Mise a jour disponible",
                f"Version actuelle: {current_label}\nNouvelle version: {latest_label}\n\n{notes}\n\nMettre a jour maintenant ?",
            )
            if do_update:
                self.perform_remote_update(self.update_info)

    def perform_remote_update(self, update_info: dict):
        url = str(update_info.get("url", "")).strip()
        latest = str(update_info.get("latest", "")).strip()
        if not url:
            messagebox.showwarning("Mise a jour", "Le lien de mise a jour est manquant dans le manifeste distant.")
            return

        if not getattr(sys, "frozen", False):
            messagebox.showinfo(
                "Mise a jour",
                "La mise a jour auto avec remplacement d'executable est disponible sur la version .exe uniquement.\n"
                "En mode script, telechargez manuellement la nouvelle version via le lien distant.",
            )
            return

        try:
            package_path = self._download_update_to_temp(url, latest or "latest")
            self._launch_external_updater(package_path)
            messagebox.showinfo("Mise a jour", "Le programme de mise a jour va s'executer. L'application va se fermer.")
            self.destroy()
        except Exception as exc:
            messagebox.showerror("Mise a jour", f"Echec de la mise a jour automatique:\n{exc}")

    def _download_update_to_temp(self, url: str, version_label: str) -> Path:
        popup = tk.Toplevel(self)
        popup.title("Telechargement mise a jour")
        popup.geometry("520x170")
        popup.transient(self)
        popup.grab_set()

        frame = ttk.Frame(popup, padding=14, style="Surface.TFrame")
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text=f"Telechargement de la version {version_label}", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        status_var = tk.StringVar(value="Preparation...")
        ttk.Label(frame, textvariable=status_var).pack(anchor="w", pady=(6, 8))
        progress = ttk.Progressbar(frame, orient="horizontal", mode="determinate", length=470, maximum=100)
        progress.pack(fill="x")
        percent_var = tk.StringVar(value="0%")
        ttk.Label(frame, textvariable=percent_var).pack(anchor="e", pady=(6, 0))

        updates_dir = Path(tempfile.gettempdir()) / "vente2_updates"
        updates_dir.mkdir(parents=True, exist_ok=True)

        parsed = urlparse(url)
        file_name = Path(parsed.path).name or f"GestionMagasinPOS_{version_label}.exe"
        out_path = updates_dir / file_name

        with urlopen(url, timeout=25) as response:
            total = int(response.headers.get("Content-Length", "0") or 0)
            if total <= 0:
                progress.configure(mode="indeterminate")
                progress.start(12)
                status_var.set("Telechargement en cours...")
                popup.update_idletasks()
                data = response.read()
                out_path.write_bytes(data)
                progress.stop()
                progress.configure(mode="determinate")
                progress["value"] = 100
                percent_var.set("100%")
            else:
                chunk_size = 64 * 1024
                downloaded = 0
                with out_path.open("wb") as f:
                    while True:
                        chunk = response.read(chunk_size)
                        if not chunk:
                            break
                        f.write(chunk)
                        downloaded += len(chunk)
                        pct = int((downloaded * 100) / total)
                        progress["value"] = pct
                        percent_var.set(f"{pct}%")
                        status_var.set(f"Telechargement... {downloaded // 1024} / {max(1, total // 1024)} Ko")
                        popup.update_idletasks()

        popup.destroy()
        return out_path

    def _launch_external_updater(self, package_path: Path):
        current_exe = Path(sys.executable).resolve()
        current_pid = os.getpid()
        updater_dir = Path(tempfile.gettempdir()) / "vente2_updates"
        updater_dir.mkdir(parents=True, exist_ok=True)
        script_path = updater_dir / f"updater_{int(time.time())}.cmd"

        script = f"""@echo off
setlocal enableextensions
set "TARGET={current_exe}"
set "SOURCE={package_path}"
set "PID={current_pid}"

:waitproc
tasklist /FI "PID eq %PID%" | find "%PID%" >nul
if not errorlevel 1 (
  timeout /t 1 /nobreak >nul
  goto waitproc
)

copy /Y "%SOURCE%" "%TARGET%" >nul
start "" "%TARGET%"
del /Q "%SOURCE%" >nul 2>&1
del /Q "%~f0" >nul 2>&1
endlocal
"""
        script_path.write_text(script, encoding="utf-8")

        flags = 0
        if hasattr(subprocess, "CREATE_NEW_PROCESS_GROUP"):
            flags |= subprocess.CREATE_NEW_PROCESS_GROUP
        if hasattr(subprocess, "DETACHED_PROCESS"):
            flags |= subprocess.DETACHED_PROCESS

        subprocess.Popen(["cmd", "/c", str(script_path)], creationflags=flags)

    def show_main(self, user: dict):
        self.current_user = user
        self.clear_container()
        MainView(self._container, self, user).pack(fill="both", expand=True)

    def setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        style.configure("App.TFrame", background=PALETTE["bg"])
        style.configure("Surface.TFrame", background=PALETTE["surface"])
        style.configure("Header.TFrame", background=PALETTE["header"])
        style.configure("Sidebar.TFrame", background=PALETTE["sidebar"])

        style.configure("TLabel", background=PALETTE["bg"], foreground=PALETTE["text_dark"], font=("Segoe UI", 10))
        style.configure("Header.TLabel", background=PALETTE["header"], foreground=PALETTE["text_light"], font=("Segoe UI", 10, "bold"))
        style.configure("Nav.TLabel", background=PALETTE["sidebar"], foreground=PALETTE["text_light"], font=("Segoe UI", 9, "bold"))
        style.configure("HeaderIcon.TLabel", background=PALETTE["header"], foreground="#d6e9ff", font=("Segoe UI", 11, "bold"))

        style.configure("TEntry", fieldbackground="#ffffff", bordercolor="#cfd9e8", lightcolor="#cfd9e8")
        style.configure("TCombobox", fieldbackground="#ffffff")

        style.configure("TButton", font=("Segoe UI", 9, "bold"), padding=(10, 6), background=PALETTE["primary"], foreground="#ffffff", borderwidth=0)
        style.map("TButton", background=[("active", "#0c59b0")], foreground=[("active", "#ffffff")])

        style.configure("Success.TButton", background=PALETTE["success"], foreground="#ffffff")
        style.map("Success.TButton", background=[("active", "#19753c")], foreground=[("active", "#ffffff")])

        style.configure("Danger.TButton", background=PALETTE["danger"], foreground="#ffffff")
        style.map("Danger.TButton", background=[("active", "#a0322b")], foreground=[("active", "#ffffff")])

        style.configure("Warning.TButton", background=PALETTE["warning"], foreground="#ffffff")
        style.map("Warning.TButton", background=[("active", "#c58514")], foreground=[("active", "#ffffff")])

        style.configure("Nav.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 10), anchor="w", background=PALETTE["sidebar"], foreground="#ffffff", borderwidth=0)
        style.map("Nav.TButton", background=[("active", "#0f5eb9")], foreground=[("active", "#ffffff")])

        style.configure("NavActive.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 10), anchor="w", background="#1170d3", foreground="#ffffff", borderwidth=0)
        style.map("NavActive.TButton", background=[("active", "#1170d3")], foreground=[("active", "#ffffff")])

        style.configure("HeaderGhost.TButton", font=("Segoe UI", 11, "bold"), padding=(8, 4), background=PALETTE["header"], foreground="#e7f3ff", borderwidth=0)
        style.map("HeaderGhost.TButton", background=[("active", "#0a4e9f")], foreground=[("active", "#ffffff")])

        style.configure("TLabelframe", background=PALETTE["surface"], bordercolor="#d8e0ef")
        style.configure("TLabelframe.Label", background=PALETTE["surface"], foreground=PALETTE["text_dark"], font=("Segoe UI", 10, "bold"))

        heading_bg = "#edf2fb" if PALETTE["bg"] != "#151c2b" else "#2b3750"
        style.configure(
            "Treeview",
            background=PALETTE["surface"],
            fieldbackground=PALETTE["surface"],
            foreground=PALETTE["text_dark"],
            rowheight=25,
            bordercolor="#d7dfeb",
        )
        style.configure("Treeview.Heading", background=heading_bg, foreground=PALETTE["text_dark"], font=("Segoe UI", 9, "bold"))

        # Hide notebook tabs; navigation is handled by the custom left sidebar.
        style.layout("Tabless.TNotebook.Tab", [])


class LoginView(ttk.Frame):
    def __init__(self, parent, app: StoreApp):
        super().__init__(parent, style="App.TFrame")
        self.app = app
        self.configure(padding=0)

        self.login_tone = {
            "sky": [PALETTE["bg"], "#c9d7ef", "#b4c7e6", "#9fb7da", "#889fc9", "#7b8fb6"],
            "mount1": PALETTE["primary"],
            "mount2": PALETTE["sidebar"],
            "mount3": "#22315a",
            "silhouette": "#161f3d",
            "card": PALETTE["header"],
            "card_text": PALETTE["text_light"],
            "card_soft": "#e8f1ff",
        }

        self.scene = tk.Canvas(self, highlightthickness=0, bg=PALETTE["bg"])
        self.scene.pack(fill="both", expand=True)
        self.scene.bind("<Configure>", self._draw_login_scene)

        self.card = tk.Frame(self.scene, bg="#ffffff", bd=0, highlightthickness=1, highlightbackground="#d8d8ef")
        self.card_window = self.scene.create_window(0, 0, window=self.card)

        tk.Label(self.card, text="Login", bg="#ffffff", fg="#ffffff", font=("Segoe UI", 28, "bold"))

        # Inner panel produces a soft glass-like container over the illustration.
        panel = tk.Frame(self.card, bg="#ffffff", bd=0)
        panel.pack(padx=1, pady=1)

        self.card_inner = tk.Frame(panel, bg=self.login_tone["card"], bd=0)
        self.card_inner.pack(fill="both", expand=True)

        tk.Label(self.card_inner, text="Login", bg=self.login_tone["card"], fg=self.login_tone["card_text"], font=("Segoe UI", 24, "bold")).pack(pady=(18, 12))

        tk.Label(self.card_inner, text="Email ID", bg=self.login_tone["card"], fg=self.login_tone["card_soft"], font=("Segoe UI", 10)).pack(anchor="w", padx=30)
        self.username = ttk.Entry(self.card_inner, width=34)
        self.username.pack(padx=30, pady=(4, 12), ipady=6)

        tk.Label(self.card_inner, text="Password", bg=self.login_tone["card"], fg=self.login_tone["card_soft"], font=("Segoe UI", 10)).pack(anchor="w", padx=30)
        self.password = ttk.Entry(self.card_inner, width=34, show="*")
        self.password.pack(padx=30, pady=(4, 8), ipady=6)

        options = tk.Frame(self.card_inner, bg=self.login_tone["card"])
        options.pack(fill="x", padx=28, pady=(4, 10))
        self.remember_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options,
            text="Remember me",
            variable=self.remember_var,
            bg=self.login_tone["card"],
            fg=self.login_tone["card_soft"],
            selectcolor=self.login_tone["card"],
            activebackground=self.login_tone["card"],
            activeforeground="#ffffff",
            font=("Segoe UI", 9),
            relief="flat",
            bd=0,
            highlightthickness=0,
        ).pack(side="left")
        tk.Label(options, text="Forgot Password?", bg=self.login_tone["card"], fg=self.login_tone["card_soft"], font=("Segoe UI", 9, "underline")).pack(side="right")

        tk.Button(
            self.card_inner,
            text="Login",
            command=self.login,
            bg="#ffffff",
            fg="#162d5a",
            activebackground="#e4edf9",
            activeforeground="#162d5a",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            bd=0,
            padx=28,
            pady=8,
            cursor="hand2",
        ).pack(fill="x", padx=28, pady=(6, 10))

        tk.Label(
            self.card_inner,
            text="Compte admin: admin / admin123",
            bg=self.login_tone["card"],
            fg=self.login_tone["card_soft"],
            font=("Segoe UI", 9),
        ).pack(pady=(2, 16))

        self._glow_offset = 0
        self._animate_login_glow()

        self.username.focus_set()
        self.bind_all("<Return>", lambda _e: self.login())

    def _draw_login_scene(self, _event=None):
        w = max(self.scene.winfo_width(), 1100)
        h = max(self.scene.winfo_height(), 700)
        c = self.scene
        c.delete("bg")

        # Sky gradient bands.
        bands = self.login_tone["sky"]
        band_h = h // len(bands)
        for i, col in enumerate(bands):
            y0 = i * band_h
            y1 = h if i == len(bands) - 1 else (i + 1) * band_h
            c.create_rectangle(0, y0, w, y1, fill=col, outline=col, tags="bg")

        # Clouds.
        c.create_polygon(0, 120, w * 0.16, 90, w * 0.32, 120, w * 0.46, 100, w * 0.62, 130, w * 0.82, 108, w, 140, w, 190, 0, 190, fill="#d8d6ea", outline="", tags="bg")
        c.create_polygon(0, 210, w * 0.2, 180, w * 0.37, 220, w * 0.55, 192, w * 0.77, 230, w, 206, w, 250, 0, 250, fill="#c4cdcf", outline="", tags="bg")

        # Mountain layers.
        c.create_polygon(0, 330, w * 0.16, 270, w * 0.29, 320, w * 0.4, 300, w * 0.55, 350, w * 0.68, 250, w * 0.84, 330, w, 280, w, h, 0, h, fill=self.login_tone["mount1"], outline="", tags="bg")
        c.create_polygon(0, 430, w * 0.12, 330, w * 0.23, 440, w * 0.35, 400, w * 0.5, 470, w * 0.64, 360, w * 0.83, 460, w, 390, w, h, 0, h, fill=self.login_tone["mount2"], outline="", tags="bg")
        c.create_polygon(0, 530, w * 0.2, 420, w * 0.32, 510, w * 0.46, 470, w * 0.63, 560, w * 0.78, 430, w * 0.93, 560, w, 520, w, h, 0, h, fill=self.login_tone["mount3"], outline="", tags="bg")

        # Foreground silhouette.
        c.create_polygon(0, h - 80, w * 0.08, h - 60, w * 0.16, h - 88, w * 0.24, h - 52, w * 0.32, h - 86, w * 0.42, h - 48, w * 0.52, h - 92, w * 0.62, h - 56, w * 0.74, h - 102, w * 0.86, h - 58, w * 0.94, h - 108, w, h - 74, w, h, 0, h, fill=self.login_tone["silhouette"], outline="", tags="bg")

        # Birds.
        c.create_text(w * 0.62, h * 0.2, text="⌒  ⌒", fill="#3f3e85", font=("Segoe UI", 18), tags="bg")

        # Card and glow.
        card_w, card_h = 420, 460
        cx, cy = w / 2, h / 2 + 10
        c.coords(self.card_window, cx, cy)
        c.create_rectangle(cx - card_w / 2 - 2, cy - card_h / 2 - 2, cx + card_w / 2 + 2, cy + card_h / 2 + 2, outline="#d6e4ff", width=2, tags="bg")
        c.create_oval(cx + 40 + self._glow_offset, cy - 140, cx + 260 + self._glow_offset, cy + 80, fill=PALETTE["primary"], outline="", stipple="gray25", tags="bg")
        self.card.configure(width=card_w, height=card_h, bg=self.login_tone["card"], highlightbackground="#d6e4ff", highlightthickness=1)

    def _animate_login_glow(self):
        self._glow_offset = 8 if self._glow_offset == 0 else 0
        self._draw_login_scene()
        self.after(1200, self._animate_login_glow)

    def login(self):
        user = self.app.service.authenticate(self.username.get().strip(), self.password.get())
        if not user:
            messagebox.showerror("Connexion", "Identifiants invalides.")
            return
        self.app.service.audit(user["id"], "Connexion", "Connexion utilisateur")
        self.app.show_main(user)


class MainView(ttk.Frame):
    def __init__(self, parent, app: StoreApp, user: dict):
        super().__init__(parent, style="App.TFrame")
        self.app = app
        self.user = user
        self.service = app.service
        self.nav_buttons: dict[str, ttk.Button] = {}
        self.active_tab_key = "dashboard"
        self._auto_collapsed_sidebar = False
        self._compact_mode = False
        self._pos_vertical_mode = False
        self.permissions = self.service.get_permissions(user["role_name"])
        self._tab_animating = False

        self.cart: list[dict] = []
        self.purchase_cart: list[dict] = []
        self.runtime_context = self.service.get_runtime_context()
        self.brand_name = self.service.get_setting("brand.name", "Magasin POS")
        self.current_theme = self.service.get_setting("ui.theme", "Clair Pro")
        self.currency_code = self.service.get_setting("app.currency", "EUR").upper()
        valid_codes = {code for code, _name in CURRENCY_CHOICES}
        if self.currency_code not in valid_codes:
            self.currency_code = "EUR"
        self.app.apply_theme(self.current_theme)
        self.icon_images: dict[str, tk.PhotoImage] = {}
        self._load_icon_images()

        self.top_bar = ttk.Frame(self, style="Header.TFrame", padding=(14, 8))
        self.top_bar.pack(fill="x")
        ttk.Button(self.top_bar, text=ICON_GLYPHS["menu"], style="HeaderGhost.TButton", command=self.toggle_sidebar).pack(side="left", padx=(0, 8))
        self.top_title = ttk.Label(self.top_bar, text=f"{self.brand_name}  •  Point de Vente", style="Header.TLabel", font=("Segoe UI", 13, "bold"))
        self.top_title.pack(side="left")

        self.top_tools = ttk.Frame(self.top_bar, style="Header.TFrame")
        self.top_tools.pack(side="right")

        self.help_icon = ttk.Label(self.top_tools, text=ICON_GLYPHS["help"], style="HeaderIcon.TLabel")
        self.help_icon.pack(side="left", padx=8)
        self.help_icon.configure(cursor="hand2")
        self.help_icon.bind("<Button-1>", lambda _e: self.show_about_dialog())
        self.bell_icon = ttk.Label(self.top_tools, text=ICON_GLYPHS["notify"], style="HeaderIcon.TLabel")
        self.bell_icon.pack(side="left", padx=8)
        self.bell_icon.bind("<Button-1>", lambda _e: self.show_notifications())
        self.badge_icon = ttk.Label(self.top_tools, text="● 8", style="HeaderIcon.TLabel")
        self.badge_icon.pack(side="left", padx=(0, 12))
        self.lbl_store_context = ttk.Label(
            self.top_tools,
            text=f"Magasin {self.runtime_context['store_id']} / Caisse {self.runtime_context['register_id']}",
            style="Header.TLabel",
        )
        self.lbl_store_context.pack(side="left", padx=(0, 12))
        self.user_label = ttk.Label(self.top_tools, text=f"{ICON_GLYPHS['users']} {user['full_name']}", style="Header.TLabel")
        self.user_label.pack(side="left", padx=(0, 6))
        self.logout_btn = ttk.Button(self.top_tools, text="Deconnexion", command=self.logout, style="Danger.TButton")
        self.logout_btn.pack(side="left", padx=(0, 12))

        body = ttk.Frame(self, style="App.TFrame")
        body.pack(fill="both", expand=True)

        self.sidebar = ttk.Frame(body, style="Sidebar.TFrame", width=230)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self.sidebar_expanded = True

        self.sidebar_header = tk.Frame(self.sidebar, bg=PALETTE["sidebar"], padx=12, pady=16)
        self.sidebar_header.pack(fill="x")

        self.sidebar_nav_wrap = tk.Frame(self.sidebar, bg=PALETTE["sidebar"])
        self.sidebar_nav_wrap.pack(fill="both", expand=True)

        self.sidebar_nav_canvas = tk.Canvas(self.sidebar_nav_wrap, bg=PALETTE["sidebar"], highlightthickness=0, bd=0)
        self.sidebar_nav_canvas.pack(side="left", fill="both", expand=True)
        self.sidebar_nav_scrollbar = ttk.Scrollbar(self.sidebar_nav_wrap, orient="vertical", command=self.sidebar_nav_canvas.yview)
        self.sidebar_nav_scrollbar.pack(side="right", fill="y")
        self.sidebar_nav_canvas.configure(yscrollcommand=self.sidebar_nav_scrollbar.set)

        self.sidebar_nav_inner = tk.Frame(self.sidebar_nav_canvas, bg=PALETTE["sidebar"])
        self.sidebar_nav_window = self.sidebar_nav_canvas.create_window((0, 0), window=self.sidebar_nav_inner, anchor="nw")
        self.sidebar_nav_inner.bind("<Configure>", self._on_sidebar_inner_configure)
        self.sidebar_nav_canvas.bind("<Configure>", self._on_sidebar_canvas_configure)
        self.sidebar_nav_canvas.bind("<MouseWheel>", self._on_sidebar_mousewheel)
        self.sidebar_nav_inner.bind("<MouseWheel>", self._on_sidebar_mousewheel)
        self.sidebar_nav_canvas.bind("<Enter>", self._bind_sidebar_wheel)
        self.sidebar_nav_canvas.bind("<Leave>", self._unbind_sidebar_wheel)
        self.sidebar_nav_inner.bind("<Enter>", self._bind_sidebar_wheel)
        self.sidebar_nav_inner.bind("<Leave>", self._unbind_sidebar_wheel)

        self.sidebar_footer = tk.Frame(self.sidebar, bg=PALETTE["sidebar"], padx=8, pady=8)
        self.sidebar_footer.pack(fill="x", side="bottom")
        self.pinned_settings_btn: ttk.Button | None = None

        content = ttk.Frame(body, style="App.TFrame", padding=(8, 8, 12, 12))
        content.pack(side="left", fill="both", expand=True)

        self.notebook = ttk.Notebook(content, style="Tabless.TNotebook")
        self.notebook.pack(fill="both", expand=True)

        self.tabs = {}
        self.create_tabs()
        self.build_sidebar_navigation()
        role_default_tab = {
            "Administrateur": "dashboard",
            "Gestionnaire": "stock",
            "Caissier": "pos",
        }
        self.show_tab(role_default_tab.get(self.user["role_name"], "dashboard"))
        self.refresh_all_data()
        self.bind("<Configure>", self._on_resize)

    def _load_icon_images(self):
        icon_dir = Path(__file__).parent / "assets" / "icons"
        for key, icon_path in resolve_icon_paths(icon_dir).items():
            try:
                image = tk.PhotoImage(file=str(icon_path))
                self.icon_images[key] = image.subsample(2, 2) if image.width() > 24 else image
            except Exception:
                continue

    def show_notifications(self):
        popup = tk.Toplevel(self)
        popup.title("Notifications")
        popup.geometry("460x360")
        popup.transient(self.winfo_toplevel())

        tree = ttk.Treeview(popup, columns=("level", "title", "msg"), show="headings", height=14)
        tree.heading("level", text="Niveau")
        tree.heading("title", text="Titre")
        tree.heading("msg", text="Message")
        tree.column("level", width=70, anchor="center")
        tree.column("title", width=120, anchor="w")
        tree.column("msg", width=240, anchor="w")
        tree.pack(fill="both", expand=True, padx=8, pady=8)

        tree.tag_configure("info", foreground=LEVEL_COLORS["info"])
        tree.tag_configure("warning", foreground=LEVEL_COLORS["warning"])
        tree.tag_configure("danger", foreground=LEVEL_COLORS["danger"])
        tree.tag_configure("success", foreground=LEVEL_COLORS["success"])

        for alert in self.service.get_recent_alerts(25):
            lvl = alert["level"] if alert["level"] in LEVEL_COLORS else "info"
            tree.insert("", "end", values=(lvl, alert["title"], alert["message"]), tags=(lvl,))

    def show_about_dialog(self):
        popup = tk.Toplevel(self)
        popup.title("A propos")
        popup.geometry("520x360")
        popup.transient(self.winfo_toplevel())
        popup.grab_set()
        popup.bind("<Escape>", lambda _e: popup.destroy())

        box = ttk.Frame(popup, padding=14, style="Surface.TFrame")
        box.pack(fill="both", expand=True)

        ttk.Label(box, text=f"{self.brand_name}", font=("Segoe UI", 14, "bold")).pack(anchor="w")
        ttk.Label(box, text="Application desktop de gestion magasin", foreground="#61758f").pack(anchor="w", pady=(2, 10))

        infos = [
            ("Version application", APP_VERSION),
            ("Theme actif", self.current_theme),
            ("Devise", self.currency_code),
            ("Utilisateur", self.user.get("full_name", "")),
            ("Role", self.user.get("role_name", "")),
            ("Contexte", f"Magasin {self.runtime_context['store_id']} / Caisse {self.runtime_context['register_id']}"),
            ("Python", sys.version.split()[0]),
            ("OS", platform.platform()),
        ]
        for key, value in infos:
            row = ttk.Frame(box, style="Surface.TFrame")
            row.pack(fill="x", pady=2)
            ttk.Label(row, text=f"{key}:", width=20).pack(side="left")
            ttk.Label(row, text=str(value)).pack(side="left")

        ttk.Separator(box).pack(fill="x", pady=10)
        ttk.Label(box, text="Support: exports, PDF, Excel, multi-role, sauvegarde/restauration", foreground="#61758f").pack(anchor="w")

        footer = ttk.Frame(box, style="Surface.TFrame")
        footer.pack(fill="x", pady=(12, 0))
        ttk.Label(footer, text="Cliquez sur Fermer pour quitter cette fenetre.", foreground="#61758f").pack(side="left")
        ttk.Button(footer, text="Fermer", width=12, command=popup.destroy).pack(side="right")

    def _pulse_badge(self, count: int):
        if not hasattr(self, "badge_icon"):
            return
        if count <= 0:
            self.badge_icon.config(foreground="#d6e9ff")
            return

        self.badge_icon.config(foreground="#ffefef")

        def back():
            if hasattr(self, "badge_icon"):
                self.badge_icon.config(foreground="#ffd5d5")

        self.after(220, back)

    def create_tabs(self):
        self.tabs["dashboard"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["dashboard"], text="Tableau de bord")
        self.build_dashboard_tab(self.tabs["dashboard"])

        self.tabs["products"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["products"], text="Produits")
        self.build_products_tab(self.tabs["products"])

        self.tabs["categories"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["categories"], text="Categories")
        self.build_categories_tab(self.tabs["categories"])

        self.tabs["pos"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["pos"], text="Ventes (POS)")
        self.build_pos_tab(self.tabs["pos"])

        self.tabs["invoices"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["invoices"], text="Facturation")
        self.build_invoices_tab(self.tabs["invoices"])

        self.tabs["stock"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["stock"], text="Stock")
        self.build_stock_tab(self.tabs["stock"])

        self.tabs["purchase"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["purchase"], text="Approvisionnement")
        self.build_purchase_tab(self.tabs["purchase"])

        self.tabs["suppliers"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["suppliers"], text="Fournisseurs")
        self.build_suppliers_tab(self.tabs["suppliers"])

        self.tabs["clients"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["clients"], text="Clients")
        self.build_clients_tab(self.tabs["clients"])

        self.tabs["reports"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["reports"], text="Rapports")
        self.build_reports_tab(self.tabs["reports"])

        self.tabs["users"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["users"], text="Utilisateurs")
        self.build_users_tab(self.tabs["users"])

        self.tabs["settings"] = ttk.Frame(self.notebook, padding=10, style="Surface.TFrame")
        self.notebook.add(self.tabs["settings"], text="Parametres")
        self.build_settings_tab(self.tabs["settings"])

    def build_sidebar_navigation(self):
        for child in self.sidebar_header.winfo_children():
            child.destroy()
        for child in self.sidebar_nav_inner.winfo_children():
            child.destroy()
        self.nav_group_labels: list[ttk.Label] = []

        logo = self.sidebar_header
        tk.Label(logo, text=f"●●  {self.brand_name}", bg=PALETTE["sidebar"], fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(anchor="w")

        groups = [
            (
                "VENTE",
                [
                    ("dashboard", "Tableau de bord"),
                    ("pos", "Point de vente"),
                    ("invoices", "Facturation"),
                    ("clients", "Clients"),
                ],
            ),
            (
                "STOCK",
                [
                    ("products", "Produits"),
                    ("categories", "Categories"),
                    ("stock", "Stock"),
                    ("purchase", "Approvisionnement"),
                    ("suppliers", "Fournisseurs"),
                ],
            ),
            (
                "PILOTAGE",
                [
                    ("reports", "Rapports"),
                    ("users", "Utilisateurs"),
                ],
            ),
        ]

        tab_permission = {
            "dashboard": "dashboard:view",
            "pos": "pos:sell",
            "products": "products:manage",
            "categories": "categories:manage",
            "stock": "stock:manage",
            "purchase": "purchase:manage",
            "suppliers": "suppliers:manage",
            "clients": "clients:manage",
            "invoices": "invoice:export",
            "reports": "reports:view",
            "users": "users:manage",
            "settings": "settings:manage",
        }

        for group_name, items in groups:
            lbl = ttk.Label(self.sidebar_nav_inner, text=group_name, style="Nav.TLabel")
            lbl.pack(anchor="w", padx=12, pady=(10, 4))
            self.nav_group_labels.append(lbl)
            for key, title in items:
                required = tab_permission.get(key)
                if required and required not in self.permissions:
                    continue
                text = f"{ICON_GLYPHS.get(key, '•')}  {title}"
                btn = ttk.Button(self.sidebar_nav_inner, text=text, style="Nav.TButton", command=lambda k=key: self.show_tab(k))
                icon_image = self.icon_images.get(key)
                if icon_image:
                    btn.configure(image=icon_image, compound="left")
                    btn._has_icon_image = True
                btn._full_text = text
                btn._short_text = ICON_GLYPHS.get(key, "•")
                btn.pack(fill="x", padx=8, pady=1)
                self.nav_buttons[key] = btn

        for child in self.sidebar_footer.winfo_children():
            child.destroy()
        self.pinned_settings_btn = None
        required = tab_permission.get("settings")
        if (not required) or (required in self.permissions):
            text = f"{ICON_GLYPHS.get('settings', '•')}  Parametres"
            btn = ttk.Button(self.sidebar_footer, text=text, style="Nav.TButton", command=lambda: self.show_tab("settings"))
            icon_image = self.icon_images.get("settings")
            if icon_image:
                btn.configure(image=icon_image, compound="left")
                btn._has_icon_image = True
            btn._full_text = text
            btn._short_text = ICON_GLYPHS.get("settings", "•")
            btn.pack(fill="x", padx=2, pady=(0, 2))
            self.pinned_settings_btn = btn
            self.nav_buttons["settings"] = btn

    def toggle_sidebar(self):
        self.sidebar_expanded = not self.sidebar_expanded
        self.sidebar.configure(width=230 if self.sidebar_expanded else 62)
        for btn in self.nav_buttons.values():
            if self.sidebar_expanded:
                btn.configure(text=getattr(btn, "_full_text", btn.cget("text")), anchor="w")
                if getattr(btn, "_has_icon_image", False):
                    btn.configure(compound="left")
            else:
                if getattr(btn, "_has_icon_image", False):
                    btn.configure(text="", compound="image", anchor="center")
                else:
                    btn.configure(text=getattr(btn, "_short_text", "•"), anchor="center")
        if self.sidebar_expanded:
            self.sidebar_footer.configure(padx=8, pady=8)
        else:
            self.sidebar_footer.configure(padx=6, pady=6)

    def _on_sidebar_inner_configure(self, _event=None):
        self.sidebar_nav_canvas.configure(scrollregion=self.sidebar_nav_canvas.bbox("all"))

    def _on_sidebar_canvas_configure(self, event):
        self.sidebar_nav_canvas.itemconfigure(self.sidebar_nav_window, width=event.width)

    def _on_sidebar_mousewheel(self, event):
        if not self.sidebar_nav_canvas.winfo_exists():
            return
        delta = getattr(event, "delta", 0)
        if delta == 0:
            return
        self.sidebar_nav_canvas.yview_scroll(int(-1 * (delta / 120)), "units")

    def _bind_sidebar_wheel(self, _event=None):
        self.bind_all("<MouseWheel>", self._on_sidebar_mousewheel)

    def _unbind_sidebar_wheel(self, _event=None):
        self.unbind_all("<MouseWheel>")

    def _on_resize(self, event=None):
        width = self.winfo_width()
        height = self.winfo_height()

        # Auto-collapse sidebar on small widths.
        if width < 1160 and self.sidebar_expanded:
            self._auto_collapsed_sidebar = True
            self.toggle_sidebar()
        elif width >= 1260 and (not self.sidebar_expanded) and self._auto_collapsed_sidebar:
            self._auto_collapsed_sidebar = False
            self.toggle_sidebar()

        # Simplify header controls for narrow layouts.
        compact = width < 1120
        very_compact = width < 1020

        if compact:
            if self.help_icon.winfo_manager():
                self.help_icon.pack_forget()
            if self.badge_icon.winfo_manager():
                self.badge_icon.pack_forget()
            if self.lbl_store_context.winfo_manager():
                self.lbl_store_context.pack_forget()
            if self.logout_btn.winfo_manager():
                self.logout_btn.pack_forget()
        else:
            if not self.help_icon.winfo_manager():
                self.help_icon.pack(side="left", padx=8, before=self.bell_icon)
            if not self.badge_icon.winfo_manager():
                self.badge_icon.pack(side="left", padx=(0, 12), before=self.user_label)
            if not self.lbl_store_context.winfo_manager():
                self.lbl_store_context.pack(side="left", padx=(0, 12), before=self.user_label)
            if not self.logout_btn.winfo_manager():
                self.logout_btn.pack(side="left", padx=(0, 12))

        if very_compact:
            self.user_label.configure(text=ICON_GLYPHS["users"])
        else:
            self.user_label.configure(text=f"{ICON_GLYPHS['users']} {self.user['full_name']}")

        self._apply_compact_mode(width < 1180)
        self._apply_sidebar_density(height < 800)
        self._apply_pos_layout(width < 1080)
        self._apply_pos_density(width < 1120)
        self._apply_table_compaction(width < 1120)
        self._arrange_dashboard_cards(width)

    def show_tab(self, key: str):
        if key not in self.tabs:
            return
        self.notebook.select(self.tabs[key])
        self.active_tab_key = key
        self._animate_tab_switch(self.tabs[key])
        if key == "pos" and hasattr(self, "pos_barcode"):
            self.pos_barcode.focus_set()
        for btn_key, button in self.nav_buttons.items():
            button.configure(style="NavActive.TButton" if btn_key == key else "Nav.TButton")

    def _animate_tab_switch(self, tab_frame: ttk.Frame):
        if self._tab_animating:
            return
        self._tab_animating = True
        try:
            base_alpha = float(self.winfo_toplevel().attributes("-alpha"))
        except Exception:
            base_alpha = 1.0

        alpha_sequence = [max(0.9, base_alpha - 0.05), base_alpha]
        padding_sequence = [18, 14, 12, 10]

        def pad_step(i: int):
            if i >= len(padding_sequence):
                return
            try:
                tab_frame.configure(padding=padding_sequence[i])
            except Exception:
                return
            self.after(30, lambda: pad_step(i + 1))

        def step(i: int):
            if i >= len(alpha_sequence):
                self._tab_animating = False
                return
            try:
                self.winfo_toplevel().attributes("-alpha", alpha_sequence[i])
            except Exception:
                self._tab_animating = False
                return
            self.after(45, lambda: step(i + 1))

        pad_step(0)
        step(0)

    def _require_permission(self, permission_key: str) -> bool:
        if permission_key in self.permissions:
            return True
        messagebox.showwarning("Permission", "Action non autorisee pour votre role.")
        return False

    def refresh_all_data(self):
        self.refresh_dashboard()
        self.refresh_products_tree()
        self.refresh_categories_tree()
        self.refresh_clients_tree()
        self.refresh_suppliers_tree()
        self.refresh_stock_views()
        self.refresh_reports()
        self.refresh_users_tree()
        self.refresh_invoices_tree()
        self.reload_pos_dropdowns()
        self.reload_purchase_dropdowns()
        if hasattr(self, "badge_icon"):
            alert_count = len(self.service.get_recent_alerts(99))
            self.badge_icon.config(text=f"● {alert_count}")
            self._pulse_badge(alert_count)
        self.show_stock_alert_popup_if_needed()

    def show_stock_alert_popup_if_needed(self):
        if "stock:manage" not in self.permissions:
            return
        alerts = self.service.stock_alerts()
        if not alerts:
            return
        top = alerts[:5]
        msg_lines = [f"- {a['name']}: stock {a['stock_qty']} / min {a['min_stock']}" for a in top]
        messagebox.showwarning("Alerte stock", "Produits critiques:\n" + "\n".join(msg_lines))

    def logout(self):
        self.service.audit(self.user["id"], "Deconnexion", "Fin de session")
        self.app.show_login()

    def build_dashboard_tab(self, tab):
        header = ttk.Frame(tab, style="Surface.TFrame")
        header.pack(fill="x", pady=(2, 8))
        ttk.Label(header, text="Pilotage en temps reel", font=("Segoe UI", 13, "bold")).pack(anchor="w")
        ttk.Label(header, text="Indicateurs clefs, tendance mensuelle et classement des produits", foreground="#5f6f84").pack(anchor="w", pady=(1, 0))

        cards = tk.Frame(tab, bg=PALETTE["surface"])
        cards.pack(fill="x", pady=(0, 10))
        self.dashboard_cards_container = cards
        self.dashboard_cards: list[tk.Frame] = []

        self.lbl_revenue, self.spark_revenue = self._metric_card(cards, "Ventes du Jour", "0.00", "#d84f3f", 0)
        self.lbl_sales_count, self.spark_sales = self._metric_card(cards, "Nombre de Ventes", "0", "#1b9e54", 1)
        self.lbl_avg_basket, self.spark_basket = self._metric_card(cards, "Panier Moyen", "0", "#2f85ff", 2)
        self.lbl_margin, self.spark_margin = self._metric_card(cards, "Marge Jour", "0", "#7e57c2", 3)
        self.lbl_low_stock, self.spark_stock = self._metric_card(cards, "Stock Faible", "0", "#e7a91a", 4)
        self._arrange_dashboard_cards(self.winfo_width())

        filters = ttk.Frame(tab, style="Surface.TFrame")
        filters.pack(fill="x", pady=(2, 10))
        ttk.Label(filters, text="Periode:").pack(side="left")
        self.dashboard_period = ttk.Combobox(filters, state="readonly", width=12)
        self.dashboard_period["values"] = ["Ce mois", "6 mois", "12 mois"]
        self.dashboard_period.set("12 mois")
        self.dashboard_period.pack(side="left", padx=6)
        self.dashboard_period.bind("<<ComboboxSelected>>", lambda _e: self.refresh_dashboard())
        ttk.Button(filters, text="Actualiser", command=self.refresh_dashboard).pack(side="left", padx=10)

        main = ttk.Frame(tab, style="Surface.TFrame")
        main.pack(fill="both", expand=True)

        chart_box = ttk.LabelFrame(main, text="Tendance des ventes", padding=10)
        chart_box.pack(side="left", fill="both", expand=True, padx=(0, 8))

        chart_bg = "#ffffff" if PALETTE["bg"] != "#151c2b" else "#1f2738"
        self.sales_chart_canvas = tk.Canvas(chart_box, bg=chart_bg, highlightthickness=0, height=330)
        self.sales_chart_canvas.pack(fill="both", expand=True)
        self.sales_chart_canvas.bind("<Configure>", lambda _e: self._draw_sales_chart())

        rank_box = ttk.LabelFrame(main, text="Classement des ventes", padding=10)
        rank_box.pack(side="left", fill="y")

        self.top_products_tree = ttk.Treeview(rank_box, columns=("nom", "vendu"), show="headings", height=14)
        self.top_products_tree.heading("nom", text="Produit")
        self.top_products_tree.heading("vendu", text="Quantite vendue")
        self.top_products_tree.column("nom", width=240, anchor="w")
        self.top_products_tree.column("vendu", width=130, anchor="e")
        self.top_products_tree.pack(fill="both", expand=True, pady=4)

    def refresh_dashboard(self):
        data = self.service.dashboard_metrics()
        self._animate_numeric_label(self.lbl_revenue, float(data["today_revenue"]), f" {self.currency_code}", decimals=2)
        self._animate_numeric_label(self.lbl_sales_count, float(data["today_sales_count"]), "", decimals=0)
        self._animate_numeric_label(self.lbl_avg_basket, float(data["avg_basket"]), f" {self.currency_code}", decimals=2)
        self._animate_numeric_label(self.lbl_margin, float(data["today_margin"]), f" {self.currency_code}", decimals=2)
        self._animate_numeric_label(self.lbl_low_stock, float(data["low_stock_count"]), "", decimals=0)

        trend_values = self._load_monthly_sales_values()
        self._draw_card_sparkline(self.spark_revenue, trend_values, "#ffffff")
        self._draw_card_sparkline(self.spark_sales, trend_values[-6:] if trend_values else [], "#ffffff")
        self._draw_card_sparkline(self.spark_basket, trend_values[-8:] if trend_values else [], "#ffffff")
        self._draw_card_sparkline(self.spark_margin, trend_values[-10:] if trend_values else [], "#ffffff")
        self._draw_card_sparkline(self.spark_stock, [max(0, 12 - v) for v in range(1, 13)], "#ffffff")
        self._draw_sales_chart()

        self.top_products_tree.delete(*self.top_products_tree.get_children())
        for row in data["top_products"]:
            self.top_products_tree.insert("", "end", values=(row["name"], row["sold"]))
        self._tree_apply_zebra(self.top_products_tree)

    def _load_monthly_sales_values(self) -> list[float]:
        period = self.dashboard_period.get() if hasattr(self, "dashboard_period") else "12 mois"
        months = 12
        if period == "Ce mois":
            months = 1
        elif period == "6 mois":
            months = 6

        rows = fetch_all(
            """
            SELECT DATE_FORMAT(created_at, '%Y-%m') AS ym,
                   IFNULL(SUM(total_amount), 0) AS total
            FROM sales
            WHERE created_at >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
            GROUP BY DATE_FORMAT(created_at, '%Y-%m')
            ORDER BY ym
            """
        )
        mapped = {r["ym"]: float(r["total"]) for r in rows}

        points = []
        y = datetime.now().year
        m = datetime.now().month
        for _ in range(months):
            key = f"{y:04d}-{m:02d}"
            points.append((key, mapped.get(key, 0.0)))
            m -= 1
            if m == 0:
                m = 12
                y -= 1
        points.reverse()
        self.dashboard_chart_labels = [p[0][5:] for p in points]
        self.dashboard_chart_values = [p[1] for p in points]
        return self.dashboard_chart_values

    def _draw_sales_chart(self):
        if not hasattr(self, "sales_chart_canvas"):
            return
        values = getattr(self, "dashboard_chart_values", [])
        labels = getattr(self, "dashboard_chart_labels", [])
        canvas = self.sales_chart_canvas
        canvas.delete("all")

        width = max(canvas.winfo_width(), 400)
        height = max(canvas.winfo_height(), 260)
        margin_l, margin_r, margin_t, margin_b = 52, 20, 20, 40
        chart_w = width - margin_l - margin_r
        chart_h = height - margin_t - margin_b

        dark = PALETTE["bg"] == "#151c2b"
        axis_color = "#6f84a1" if dark else "#9fb3cd"
        text_color = "#bdd0e8" if dark else "#667a95"
        bar_fill = "#4aa2ff" if dark else "#3f93f2"
        bar_outline = "#2d82de" if dark else "#2a79d1"

        canvas.create_line(margin_l, margin_t, margin_l, height - margin_b, fill=axis_color, width=1)
        canvas.create_line(margin_l, height - margin_b, width - margin_r, height - margin_b, fill=axis_color, width=1)

        if not values:
            canvas.create_text(width // 2, height // 2, text="Aucune vente pour la periode", fill=text_color, font=("Segoe UI", 11))
            return

        max_v = max(values) if max(values) > 0 else 1.0
        gap = chart_w / max(len(values), 1)
        bar_w = max(18, int(gap * 0.58))

        for i, v in enumerate(values):
            x0 = margin_l + i * gap + (gap - bar_w) / 2
            y1 = height - margin_b
            bar_h = (v / max_v) * chart_h
            y0 = y1 - bar_h
            canvas.create_rectangle(x0, y0, x0 + bar_w, y1, fill=bar_fill, outline=bar_outline)
            if len(values) <= 8:
                canvas.create_text(x0 + bar_w / 2, y0 - 10, text=f"{v:.0f}", fill=text_color, font=("Segoe UI", 8))
            if i < len(labels):
                canvas.create_text(x0 + bar_w / 2, y1 + 12, text=labels[i], fill=text_color, font=("Segoe UI", 8))

        for t in range(0, 6):
            y = margin_t + (chart_h * t / 5)
            value = max_v * (1 - t / 5)
            canvas.create_text(margin_l - 8, y, text=f"{value:.0f}", fill=text_color, font=("Segoe UI", 8), anchor="e")

    def _draw_card_sparkline(self, canvas: tk.Canvas, values: list[float], color: str):
        canvas.delete("all")
        w, h = 168, 38
        canvas.create_rectangle(0, 0, w, h, fill=canvas["bg"], outline=canvas["bg"])
        if not values:
            return
        max_v = max(values) if max(values) > 0 else 1.0
        step = w / max(len(values) - 1, 1)
        points = []
        for i, v in enumerate(values):
            x = i * step
            y = h - ((v / max_v) * (h - 6)) - 2
            points.extend([x, y])
        if len(points) >= 4:
            canvas.create_line(*points, fill=color, width=2, smooth=True)

    def build_products_tab(self, tab):
        form = ttk.LabelFrame(tab, text="Produit", padding=10)
        form.pack(fill="x")

        self.p_name = self._entry(form, "Nom", 0, 0)
        self.p_barcode = self._entry(form, "Code-barres", 0, 2)
        ttk.Button(form, text="Generer", command=lambda: self.p_barcode.insert(0, generate_barcode())).grid(row=1, column=4, padx=6)

        self.p_brand = self._entry(form, "Marque", 2, 0)
        self.p_purchase = self._entry(form, "Prix achat", 2, 2)
        self.p_sale = self._entry(form, "Prix vente", 2, 4)

        self.p_stock = self._entry(form, "Stock", 4, 0)
        self.p_min = self._entry(form, "Stock min", 4, 2)

        ttk.Label(form, text="Categorie").grid(row=4, column=4, sticky="w", pady=3)
        self.category_combo = ttk.Combobox(form, state="readonly", width=25)
        self.category_combo.grid(row=5, column=4, sticky="w")

        actions = ttk.Frame(form)
        actions.grid(row=6, column=0, columnspan=5, sticky="w", pady=8)
        ttk.Button(actions, text="Ajouter", command=self.add_product).pack(side="left", padx=4)
        ttk.Button(actions, text="Modifier", command=self.update_product).pack(side="left", padx=4)
        ttk.Button(actions, text="Supprimer", command=self.delete_product).pack(side="left", padx=4)
        ttk.Button(actions, text="Vider", command=self.clear_product_form).pack(side="left", padx=4)

        search_box = ttk.Frame(tab)
        search_box.pack(fill="x", pady=8)
        ttk.Label(search_box, text="Recherche").pack(side="left")
        self.product_search = ttk.Entry(search_box, width=40)
        self.product_search.pack(side="left", padx=5)
        self.product_search.bind("<KeyRelease>", lambda _e: self.refresh_products_tree())
        ttk.Button(search_box, text="Importer Excel", command=self.import_products_excel).pack(side="left", padx=4)
        ttk.Button(search_box, text="Exporter Excel", command=self.export_products_excel).pack(side="left", padx=4)

        cols = ("id", "nom", "barcode", "categorie", "achat", "vente", "marque", "stock", "min")
        self.products_tree = ttk.Treeview(tab, columns=cols, show="headings", height=12)
        for c, title in [
            ("id", "ID"),
            ("nom", "Nom"),
            ("barcode", "Code"),
            ("categorie", "Categorie"),
            ("achat", "Achat"),
            ("vente", "Vente"),
            ("marque", "Marque"),
            ("stock", "Stock"),
            ("min", "Min"),
        ]:
            self.products_tree.heading(c, text=title)
        self.products_tree.pack(fill="both", expand=True)
        self.products_tree.bind("<<TreeviewSelect>>", self.on_product_select)

    def refresh_products_tree(self):
        categories = self.service.list_categories()
        self.category_by_name = {c["name"]: c["id"] for c in categories}
        self.category_combo["values"] = [c["name"] for c in categories]

        self.products_tree.delete(*self.products_tree.get_children())
        rows = self.service.list_products(self.product_search.get() if hasattr(self, "product_search") else "")
        for row in rows:
            self.products_tree.insert(
                "",
                "end",
                values=(
                    row["id"],
                    row["name"],
                    row["barcode"],
                    row["category"],
                    row["purchase_price"],
                    row["sale_price"],
                    row["brand"],
                    row["stock_qty"],
                    row["min_stock"],
                ),
            )
        self._tree_apply_zebra(self.products_tree)

    def add_product(self):
        if not self._require_permission("products:manage"):
            return
        try:
            category_id = self.category_by_name.get(self.category_combo.get())
            self.service.add_product(
                self.p_name.get().strip(),
                self.p_barcode.get().strip(),
                category_id,
                float(self.p_purchase.get() or 0),
                float(self.p_sale.get() or 0),
                self.p_brand.get().strip(),
                int(self.p_stock.get() or 0),
                int(self.p_min.get() or 0),
            )
            self.service.audit(self.user["id"], "Produit ajoute", self.p_name.get().strip())
            self.clear_product_form()
            self.refresh_products_tree()
            self.reload_pos_dropdowns()
            messagebox.showinfo("Produit", "Produit ajoute avec succes.")
        except Exception as exc:
            messagebox.showerror("Produit", str(exc))

    def update_product(self):
        if not self._require_permission("products:manage"):
            return
        item = self.products_tree.selection()
        if not item:
            return
        try:
            product_id = int(self.products_tree.item(item[0], "values")[0])
            category_id = self.category_by_name.get(self.category_combo.get())
            self.service.update_product(
                product_id,
                self.p_name.get().strip(),
                self.p_barcode.get().strip(),
                category_id,
                float(self.p_purchase.get() or 0),
                float(self.p_sale.get() or 0),
                self.p_brand.get().strip(),
                int(self.p_stock.get() or 0),
                int(self.p_min.get() or 0),
            )
            self.service.audit(self.user["id"], "Produit modifie", f"ID {product_id}")
            self.refresh_products_tree()
            self.reload_pos_dropdowns()
            messagebox.showinfo("Produit", "Produit modifie.")
        except Exception as exc:
            messagebox.showerror("Produit", str(exc))

    def delete_product(self):
        if not self._require_permission("products:manage"):
            return
        item = self.products_tree.selection()
        if not item:
            return
        product_id = int(self.products_tree.item(item[0], "values")[0])
        if not messagebox.askyesno("Produit", "Supprimer ce produit ?"):
            return
        try:
            self.service.delete_product(product_id)
            self.service.audit(self.user["id"], "Produit supprime", f"ID {product_id}")
            self.refresh_products_tree()
            self.reload_pos_dropdowns()
        except Exception as exc:
            messagebox.showerror("Produit", str(exc))

    def on_product_select(self, _event=None):
        sel = self.products_tree.selection()
        if not sel:
            return
        vals = self.products_tree.item(sel[0], "values")
        self._set_entry(self.p_name, vals[1])
        self._set_entry(self.p_barcode, vals[2])
        self.category_combo.set(vals[3])
        self._set_entry(self.p_purchase, vals[4])
        self._set_entry(self.p_sale, vals[5])
        self._set_entry(self.p_brand, vals[6])
        self._set_entry(self.p_stock, vals[7])
        self._set_entry(self.p_min, vals[8])

    def clear_product_form(self):
        for e in [self.p_name, self.p_barcode, self.p_brand, self.p_purchase, self.p_sale, self.p_stock, self.p_min]:
            self._set_entry(e, "")
        self.category_combo.set("")

    def export_products_excel(self):
        if not self._require_permission("products:manage"):
            return
        rows = self.service.list_products("")
        wb = Workbook()
        ws = wb.active
        ws.title = "Produits"
        ws.append(["Nom", "Code-barres", "Categorie", "Prix achat", "Prix vente", "Marque", "Stock", "Stock min"])
        for r in rows:
            ws.append([r["name"], r["barcode"], r["category"], r["purchase_price"], r["sale_price"], r["brand"], r["stock_qty"], r["min_stock"]])
        out_dir = EXPORTS_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / f"produits_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(path)
        messagebox.showinfo("Produits", f"Export Excel cree: {path}")

    def import_products_excel(self):
        if not self._require_permission("products:manage"):
            return
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        wb = load_workbook(path)
        ws = wb.active
        imported = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0] or not row[1]:
                continue
            category_name = str(row[2] or "Divers").strip()
            cat = next((x for x in self.service.list_categories() if x["name"].lower() == category_name.lower()), None)
            cat_id = cat["id"] if cat else self.service.add_category(category_name)
            try:
                self.service.add_product(
                    str(row[0]).strip(),
                    str(row[1]).strip(),
                    cat_id,
                    float(row[3] or 0),
                    float(row[4] or 0),
                    str(row[5] or "").strip(),
                    int(row[6] or 0),
                    int(row[7] or 0),
                )
                imported += 1
            except Exception:
                # Ignore duplicate lines and continue import.
                continue
        self.refresh_products_tree()
        self.reload_pos_dropdowns()
        messagebox.showinfo("Produits", f"Import termine: {imported} produit(s) ajoute(s).")

    def build_categories_tab(self, tab):
        row = ttk.Frame(tab)
        row.pack(fill="x")
        ttk.Label(row, text="Nom categorie").pack(side="left")
        self.cat_name = ttk.Entry(row, width=40)
        self.cat_name.pack(side="left", padx=5)
        ttk.Button(row, text="Ajouter", command=self.add_category).pack(side="left")

        self.categories_tree = ttk.Treeview(tab, columns=("id", "name"), show="headings", height=15)
        self.categories_tree.heading("id", text="ID")
        self.categories_tree.heading("name", text="Nom")
        self.categories_tree.pack(fill="both", expand=True, pady=8)

    def add_category(self):
        if not self._require_permission("categories:manage"):
            return
        name = self.cat_name.get().strip()
        if not name:
            return
        try:
            self.service.add_category(name)
            self.cat_name.delete(0, "end")
            self.refresh_categories_tree()
            self.refresh_products_tree()
        except Exception as exc:
            messagebox.showerror("Categories", str(exc))

    def refresh_categories_tree(self):
        self.categories_tree.delete(*self.categories_tree.get_children())
        for row in self.service.list_categories():
            self.categories_tree.insert("", "end", values=(row["id"], row["name"]))
        self._tree_apply_zebra(self.categories_tree)

    def build_pos_tab(self, tab):
        header = ttk.Frame(tab, style="Surface.TFrame")
        header.pack(fill="x", pady=(2, 8))
        ttk.Label(header, text="Encaissement rapide", font=("Segoe UI", 13, "bold")).pack(anchor="w")
        ttk.Label(header, text="Scan produit, verification panier et validation paiement", foreground="#5f6f84").pack(anchor="w", pady=(1, 0))

        container = ttk.Frame(tab, style="Surface.TFrame")
        container.pack(fill="both", expand=True)
        self.pos_container = container

        left = ttk.Frame(container, style="Surface.TFrame")
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))
        self.pos_left = left

        right = ttk.Frame(container, style="Surface.TFrame")
        right.pack(side="left", fill="y")
        right.configure(width=320)
        right.pack_propagate(False)
        self.pos_right = right

        head = ttk.LabelFrame(left, text="Scanner code-barres", padding=8)
        head.pack(fill="x")
        self.pos_head = head
        ttk.Label(head, text="Code-barres").pack(side="left")
        self.pos_barcode = ttk.Entry(head, width=30)
        self.pos_barcode.pack(side="left", padx=6)
        self.pos_barcode.bind("<Return>", lambda _e: self.scan_barcode())
        ttk.Button(head, text="Scanner", command=self.scan_barcode).pack(side="left")

        ttk.Label(head, text="Produit").pack(side="left", padx=(20, 4))
        self.pos_product_combo = ttk.Combobox(head, state="readonly", width=30)
        self.pos_product_combo.pack(side="left", padx=4)
        ttk.Button(head, text="Ajouter produit", command=self.add_selected_product_to_cart).pack(side="left")

        cols = ("id", "name", "qty", "price", "total")
        self.cart_tree = ttk.Treeview(left, columns=cols, show="headings", height=18)
        for c, t in [("id", "ID"), ("name", "Produit"), ("qty", "Qte"), ("price", "Prix"), ("total", "Total")]:
            self.cart_tree.heading(c, text=t)
        self.cart_tree.pack(fill="both", expand=True, pady=10)

        actions = ttk.Frame(left, style="Surface.TFrame")
        actions.pack(fill="x")
        self.pos_actions = actions
        ttk.Button(actions, text="+ Quantite", command=lambda: self.change_cart_qty(1)).pack(side="left", padx=4, pady=(0, 2))
        ttk.Button(actions, text="- Quantite", command=lambda: self.change_cart_qty(-1)).pack(side="left", padx=4, pady=(0, 2))
        ttk.Button(actions, text="Retirer ligne", command=self.remove_cart_item, style="Danger.TButton").pack(side="left", padx=4, pady=(0, 2))
        ttk.Button(actions, text="Vider panier", command=self.clear_cart, style="Warning.TButton").pack(side="left", padx=4, pady=(0, 2))

        summary = ttk.LabelFrame(right, text="Resume de la vente", padding=10)
        summary.pack(fill="x", pady=(0, 8))
        self.pos_summary = summary

        ttk.Label(summary, text="Client").grid(row=0, column=0, sticky="w")
        self.pos_client_combo = ttk.Combobox(summary, state="readonly", width=24)
        self.pos_client_combo.grid(row=1, column=0, padx=4, pady=(0, 6), sticky="w")

        ttk.Label(summary, text="Remise").grid(row=2, column=0, sticky="w")
        self.pos_discount = ttk.Entry(summary, width=14)
        self.pos_discount.insert(0, "0")
        self.pos_discount.grid(row=3, column=0, padx=4, pady=(0, 6), sticky="w")

        ttk.Label(summary, text="TVA (%)").grid(row=4, column=0, sticky="w")
        self.pos_vat = ttk.Entry(summary, width=14)
        self.pos_vat.insert(0, "18")
        self.pos_vat.grid(row=5, column=0, padx=4, pady=(0, 8), sticky="w")

        ttk.Separator(summary, orient="horizontal").grid(row=6, column=0, sticky="ew", pady=6)
        self.lbl_cart_total = ttk.Label(summary, text=f"Total: {self._money(0)}", font=("Segoe UI", 16, "bold"))
        self.lbl_cart_total.grid(row=7, column=0, sticky="w", padx=4)

        pay = ttk.LabelFrame(right, text="Paiement", padding=10)
        pay.pack(fill="x")
        self.pos_pay = pay

        ttk.Label(pay, text="Mode de paiement").pack(anchor="w")
        self.pos_payment = ttk.Combobox(pay, state="readonly", width=24)
        self.pos_payment["values"] = ["Especes", "Mobile Money", "Carte bancaire"]
        self.pos_payment.set("Especes")
        self.pos_payment.pack(anchor="w", pady=4)

        ttk.Button(pay, text="Especes", command=lambda: self.pos_payment.set("Especes")).pack(fill="x", pady=2)
        ttk.Button(pay, text="Carte bancaire", command=lambda: self.pos_payment.set("Carte bancaire")).pack(fill="x", pady=2)
        ttk.Button(pay, text="Mobile Money", command=lambda: self.pos_payment.set("Mobile Money"), style="Success.TButton").pack(fill="x", pady=2)
        ttk.Button(pay, text="Valider paiement", command=self.finalize_sale, style="Success.TButton").pack(fill="x", pady=(10, 2))

        self._apply_pos_layout(self.winfo_width() < 1080)
        self._apply_pos_density(self.winfo_width() < 1120)

    def reload_pos_dropdowns(self):
        products = self.service.list_products("")
        self.pos_products_by_label = {
            f"{p['name']} ({p['barcode']})": p for p in products
        }
        if hasattr(self, "pos_product_combo"):
            self.pos_product_combo["values"] = list(self.pos_products_by_label.keys())

        clients = self.service.list_clients()
        self.clients_by_name = {c["full_name"]: c for c in clients}
        if hasattr(self, "pos_client_combo"):
            self.pos_client_combo["values"] = ["(Sans client)"] + [c["full_name"] for c in clients]
            self.pos_client_combo.set("(Sans client)")

    def scan_barcode(self):
        barcode = self.pos_barcode.get().strip()
        if not barcode:
            self.pos_barcode.focus_set()
            return
        product = self.service.find_product_by_barcode(barcode)
        if not product:
            messagebox.showwarning("POS", "Produit introuvable pour ce code-barres.")
            self.pos_barcode.focus_set()
            return
        self.add_to_cart(product)
        self.bell()
        self.pos_barcode.delete(0, "end")
        self.pos_barcode.focus_set()

    def add_selected_product_to_cart(self):
        label = self.pos_product_combo.get()
        product = self.pos_products_by_label.get(label)
        if not product:
            return
        self.add_to_cart(product)

    def add_to_cart(self, product: dict):
        for item in self.cart:
            if item["product_id"] == product["id"]:
                item["qty"] += 1
                self.render_cart()
                return
        self.cart.append(
            {
                "product_id": product["id"],
                "name": product["name"],
                "qty": 1,
                "unit_price": float(product["sale_price"]),
            }
        )
        self.render_cart()

    def change_cart_qty(self, delta: int):
        sel = self.cart_tree.selection()
        if not sel:
            return
        pid = int(self.cart_tree.item(sel[0], "values")[0])
        for item in self.cart:
            if item["product_id"] == pid:
                item["qty"] = max(1, item["qty"] + delta)
                break
        self.render_cart()

    def remove_cart_item(self):
        sel = self.cart_tree.selection()
        if not sel:
            return
        pid = int(self.cart_tree.item(sel[0], "values")[0])
        self.cart = [x for x in self.cart if x["product_id"] != pid]
        self.render_cart()

    def clear_cart(self):
        self.cart = []
        self.render_cart()

    def render_cart(self):
        self.cart_tree.delete(*self.cart_tree.get_children())
        total = 0.0
        for item in self.cart:
            line = item["qty"] * item["unit_price"]
            total += line
            self.cart_tree.insert("", "end", values=(item["product_id"], item["name"], item["qty"], item["unit_price"], line))
        self.lbl_cart_total.config(text=f"Total: {self._money(total)}")
        self._tree_apply_zebra(self.cart_tree)

    def finalize_sale(self):
        if not self._require_permission("pos:sell"):
            return
        try:
            if not self.cart:
                messagebox.showwarning("POS", "Panier vide.")
                return
            discount = float(self.pos_discount.get() or 0)
            vat_rate = (float(self.pos_vat.get() or 0)) / 100
            client_name = self.pos_client_combo.get()
            client_id = None if client_name in ("", "(Sans client)") else self.clients_by_name[client_name]["id"]
            invoice = self.service.create_sale(
                user_id=self.user["id"],
                client_id=client_id,
                payment_mode=self.pos_payment.get(),
                discount_amount=discount,
                vat_rate=vat_rate,
                items=self.cart,
            )
            self.service.audit(self.user["id"], "Vente", invoice)
            pdf_path = self.generate_invoice_pdf(invoice)
            ticket_path = self.generate_ticket_80mm_pdf(invoice)
            self.clear_cart()
            self.refresh_all_data()
            messagebox.showinfo("POS", f"Vente enregistree: {invoice}\nFacture PDF: {pdf_path}\nTicket 80mm: {ticket_path}")
            self.pos_barcode.focus_set()
        except Exception as exc:
            messagebox.showerror("POS", str(exc))

    def generate_invoice_pdf(self, invoice_number: str) -> str:
        rows = fetch_all(
            """
            SELECT s.invoice_number, s.created_at, s.payment_mode, s.discount_amount, s.vat_amount, s.total_amount,
                   IFNULL(c.full_name, 'Client comptoir') AS client_name,
                   p.name AS product_name, si.qty, si.unit_price, si.line_total
            FROM sales s
            LEFT JOIN clients c ON c.id = s.client_id
            JOIN sale_items si ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            WHERE s.invoice_number = %s
            """,
            (invoice_number,),
        )
        if not rows:
            raise ValueError("Facture introuvable.")

        out_dir = EXPORTS_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{invoice_number}.pdf"

        c = canvas.Canvas(str(out_path), pagesize=A4)
        width, height = A4
        y = height - 40

        primary = colors.HexColor(PALETTE["primary"])
        surface = colors.HexColor("#f6f9ff")
        table_header = colors.HexColor("#e8f0ff")

        c.setFillColor(colors.HexColor(PALETTE["primary"]))
        c.rect(0, height - 62, width, 62, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, height - 38, f"Facture de vente - {self.brand_name}")
        c.setFont("Helvetica", 9)
        c.drawString(40, height - 52, f"Version {APP_VERSION}  •  Ticket: {invoice_number}")
        y -= 20

        header = rows[0]
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 10)
        c.drawString(40, y, f"Numero: {header['invoice_number']}")
        y -= 14
        c.drawString(40, y, f"Date: {header['created_at']}")
        y -= 14
        c.drawString(40, y, f"Client: {header['client_name']}")
        y -= 14
        c.drawString(40, y, "Paiement:")
        pay_mode = str(header["payment_mode"])
        badge_color = colors.HexColor("#34a853") if "Mobile" in pay_mode else (colors.HexColor("#1a73e8") if "Carte" in pay_mode else colors.HexColor("#5f6368"))
        c.setFillColor(badge_color)
        c.roundRect(96, y - 4, 86, 14, 5, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(102, y, pay_mode[:16])
        y -= 24

        c.setFillColor(surface)
        c.roundRect(34, y - 8, width - 68, 22 + (len(rows) * 13), 6, fill=1, stroke=0)
        c.setFillColor(table_header)
        c.rect(36, y - 5, width - 72, 18, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.HexColor("#183153"))
        c.drawString(40, y, "Produit")
        c.drawString(290, y, "Qte")
        c.drawString(340, y, "P.U")
        c.drawString(420, y, "Total")
        y -= 14

        c.setFont("Helvetica", 10)
        for idx, line in enumerate(rows):
            if y < 70:
                c.showPage()
                y = height - 40
            if idx % 2 == 0:
                c.setFillColor(colors.HexColor("#f0f6ff"))
                c.rect(36, y - 3, width - 72, 13, fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#1e2f45"))
            c.drawString(40, y, str(line["product_name"])[:42])
            c.drawString(290, y, str(line["qty"]))
            c.drawString(340, y, f"{float(line['unit_price']):.2f}")
            c.drawString(420, y, f"{float(line['line_total']):.2f}")
            y -= 13

        y -= 16
        c.setFillColor(primary)
        c.roundRect(332, y - 32, 190, 54, 6, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(340, y, f"Remise: {self._money(float(header['discount_amount']))}")
        y -= 14
        c.drawString(340, y, f"TVA: {self._money(float(header['vat_amount']))}")
        y -= 14
        c.drawString(340, y, f"Total net: {self._money(float(header['total_amount']))}")
        c.save()
        return str(out_path)

    def generate_ticket_80mm_pdf(self, invoice_number: str) -> str:
        rows = fetch_all(
            """
            SELECT s.invoice_number, s.created_at, s.payment_mode, s.discount_amount, s.vat_amount, s.total_amount,
                   IFNULL(c.full_name, 'Client comptoir') AS client_name,
                   p.name AS product_name, si.qty, si.unit_price, si.line_total
            FROM sales s
            LEFT JOIN clients c ON c.id = s.client_id
            JOIN sale_items si ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            WHERE s.invoice_number = %s
            """,
            (invoice_number,),
        )
        if not rows:
            raise ValueError("Facture introuvable.")

        width = 226.77  # 80mm
        height = max(420, 170 + (len(rows) * 16))

        out_dir = EXPORTS_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{invoice_number}_ticket80mm.pdf"

        c = canvas.Canvas(str(out_path), pagesize=(width, height))
        y = height - 22
        c.setFillColor(colors.HexColor(PALETTE["primary"]))
        c.rect(0, height - 34, width, 34, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width / 2, y, self.brand_name[:24])
        y -= 14
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 8)
        header = rows[0]
        c.drawString(8, y, f"Facture: {header['invoice_number']}")
        y -= 11
        c.drawString(8, y, f"Date: {header['created_at']}")
        y -= 11
        c.drawString(8, y, f"Client: {str(header['client_name'])[:22]}")
        y -= 10
        c.line(8, y, width - 8, y)
        y -= 10

        c.setFont("Helvetica-Bold", 8)
        c.drawString(8, y, "Article")
        c.drawString(126, y, "Qte")
        c.drawRightString(width - 8, y, "Total")
        y -= 10
        c.setFont("Helvetica", 8)

        for line in rows:
            c.drawString(8, y, str(line["product_name"])[:20])
            c.drawString(126, y, str(line["qty"]))
            c.drawRightString(width - 8, y, self._money(float(line['line_total'])))
            y -= 11

        c.line(8, y, width - 8, y)
        y -= 11
        c.drawString(8, y, "Remise")
        c.drawRightString(width - 8, y, self._money(float(header['discount_amount'])))
        y -= 10
        c.drawString(8, y, "TVA")
        c.drawRightString(width - 8, y, self._money(float(header['vat_amount'])))
        y -= 10
        c.setFont("Helvetica-Bold", 9)
        c.drawString(8, y, "TOTAL")
        c.drawRightString(width - 8, y, self._money(float(header['total_amount'])))
        y -= 18
        c.setFont("Helvetica", 8)
        c.drawCentredString(width / 2, y, "Merci de votre visite")
        c.save()
        return str(out_path)

    def build_invoices_tab(self, tab):
        search_box = ttk.Frame(tab)
        search_box.pack(fill="x", pady=(0, 6))
        ttk.Label(search_box, text="Recherche facture").pack(side="left")
        self.invoices_search = ttk.Entry(search_box, width=34)
        self.invoices_search.pack(side="left", padx=6)
        self.invoices_search.bind("<KeyRelease>", lambda _e: self.refresh_invoices_tree())

        self.invoices_tree = ttk.Treeview(
            tab,
            columns=("invoice", "total", "payment", "date"),
            show="headings",
            height=17,
        )
        for c, t in [("invoice", "Facture"), ("total", "Total"), ("payment", "Paiement"), ("date", "Date")]:
            self.invoices_tree.heading(c, text=t)
        self.invoices_tree.pack(fill="both", expand=True)

        row = ttk.Frame(tab)
        row.pack(fill="x", pady=8)
        ttk.Button(row, text="Actualiser", command=self.refresh_invoices_tree).pack(side="left", padx=3)
        ttk.Button(row, text="Exporter PDF", command=self.export_selected_invoice).pack(side="left", padx=3)

        ret = ttk.LabelFrame(tab, text="Retour produit", padding=8)
        ret.pack(fill="x", pady=6)
        ttk.Label(ret, text="Facture").grid(row=0, column=0)
        self.ret_invoice = ttk.Entry(ret, width=18)
        self.ret_invoice.grid(row=1, column=0, padx=4)
        ttk.Label(ret, text="Produit ID").grid(row=0, column=1)
        self.ret_product_id = ttk.Entry(ret, width=10)
        self.ret_product_id.grid(row=1, column=1, padx=4)
        ttk.Label(ret, text="Quantite").grid(row=0, column=2)
        self.ret_qty = ttk.Entry(ret, width=10)
        self.ret_qty.grid(row=1, column=2, padx=4)
        ttk.Label(ret, text="Motif").grid(row=0, column=3)
        self.ret_reason = ttk.Entry(ret, width=40)
        self.ret_reason.grid(row=1, column=3, padx=4)
        ttk.Button(ret, text="Valider retour", command=self.register_return).grid(row=1, column=4, padx=6)

    def refresh_invoices_tree(self):
        self.invoices_tree.delete(*self.invoices_tree.get_children())
        term = self.invoices_search.get().strip().lower() if hasattr(self, "invoices_search") else ""
        for row in self.service.recent_sales():
            if term:
                blob = f"{row['invoice_number']} {row['payment_mode']} {row['created_at']}".lower()
                if term not in blob:
                    continue
            self.invoices_tree.insert("", "end", values=(row["invoice_number"], row["total_amount"], row["payment_mode"], row["created_at"]))
        self._tree_apply_zebra(self.invoices_tree)

    def export_selected_invoice(self):
        if not self._require_permission("invoice:export"):
            return
        sel = self.invoices_tree.selection()
        if not sel:
            return
        invoice = self.invoices_tree.item(sel[0], "values")[0]
        try:
            path = self.generate_invoice_pdf(invoice)
            messagebox.showinfo("Facturation", f"Facture exportee: {path}")
        except Exception as exc:
            messagebox.showerror("Facturation", str(exc))

    def register_return(self):
        if not self._require_permission("stock:manage"):
            return
        try:
            self.service.register_return(
                self.ret_invoice.get().strip(),
                int(self.ret_product_id.get().strip()),
                int(self.ret_qty.get().strip()),
                self.ret_reason.get().strip(),
            )
            self.service.audit(self.user["id"], "Retour produit", self.ret_invoice.get().strip())
            self.refresh_all_data()
            messagebox.showinfo("Retour", "Retour enregistre.")
        except Exception as exc:
            messagebox.showerror("Retour", str(exc))

    def build_stock_tab(self, tab):
        ttk.Label(tab, text="Produits en stock faible", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        self.stock_alert_tree = ttk.Treeview(tab, columns=("id", "name", "barcode", "stock", "min", "etat"), show="headings", height=8)
        for c, t in [("id", "ID"), ("name", "Produit"), ("barcode", "Code"), ("stock", "Stock"), ("min", "Min"), ("etat", "Statut")]:
            self.stock_alert_tree.heading(c, text=t)
        self.stock_alert_tree.pack(fill="x", pady=6)

        ttk.Label(tab, text="Historique des mouvements", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(10, 0))
        search_row = ttk.Frame(tab)
        search_row.pack(fill="x", pady=(4, 0))
        ttk.Label(search_row, text="Recherche mouvements").pack(side="left")
        self.stock_search = ttk.Entry(search_row, width=34)
        self.stock_search.pack(side="left", padx=6)
        self.stock_search.bind("<KeyRelease>", lambda _e: self.refresh_stock_views())

        self.stock_move_tree = ttk.Treeview(tab, columns=("date", "product", "type", "qty", "note"), show="headings", height=10)
        for c, t in [("date", "Date"), ("product", "Produit"), ("type", "Type"), ("qty", "Variation"), ("note", "Note")]:
            self.stock_move_tree.heading(c, text=t)
        self.stock_move_tree.pack(fill="both", expand=True, pady=6)

        ttk.Button(tab, text="Actualiser", command=self.refresh_stock_views).pack(anchor="w")

    def refresh_stock_views(self):
        self.stock_alert_tree.delete(*self.stock_alert_tree.get_children())
        for row in self.service.stock_alerts():
            status = "Rupture" if int(row["stock_qty"]) <= 0 else "Stock faible"
            tag = "danger" if status == "Rupture" else "warning"
            self.stock_alert_tree.insert("", "end", values=(row["id"], row["name"], row["barcode"], row["stock_qty"], row["min_stock"], status), tags=(tag,))
        self._tree_apply_zebra(self.stock_alert_tree)
        self.stock_alert_tree.tag_configure("danger", foreground="#b00020")
        self.stock_alert_tree.tag_configure("warning", foreground="#a06100")

        self.stock_move_tree.delete(*self.stock_move_tree.get_children())
        term = self.stock_search.get().strip().lower() if hasattr(self, "stock_search") else ""
        for row in self.service.stock_movements():
            if term:
                blob = f"{row['created_at']} {row['product_name']} {row['movement_type']} {row['note']}".lower()
                if term not in blob:
                    continue
            self.stock_move_tree.insert(
                "",
                "end",
                values=(row["created_at"], row["product_name"], row["movement_type"], row["qty_change"], row["note"]),
            )
        self._tree_apply_zebra(self.stock_move_tree)

    def build_purchase_tab(self, tab):
        form = ttk.LabelFrame(tab, text="Reception approvisionnement", padding=8)
        form.pack(fill="x")

        ttk.Label(form, text="Fournisseur").grid(row=0, column=0, sticky="w")
        self.purchase_supplier_combo = ttk.Combobox(form, state="readonly", width=35)
        self.purchase_supplier_combo.grid(row=1, column=0, padx=4)

        ttk.Label(form, text="Date livraison (YYYY-MM-DD)").grid(row=0, column=1, sticky="w")
        self.purchase_date = ttk.Entry(form, width=16)
        self.purchase_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.purchase_date.grid(row=1, column=1, padx=4)

        ttk.Label(form, text="Produit").grid(row=2, column=0, sticky="w")
        self.purchase_product_combo = ttk.Combobox(form, state="readonly", width=35)
        self.purchase_product_combo.grid(row=3, column=0, padx=4)

        ttk.Label(form, text="Quantite").grid(row=2, column=1, sticky="w")
        self.purchase_qty = ttk.Entry(form, width=12)
        self.purchase_qty.grid(row=3, column=1, padx=4, sticky="w")

        ttk.Label(form, text="Cout unitaire").grid(row=2, column=2, sticky="w")
        self.purchase_cost = ttk.Entry(form, width=12)
        self.purchase_cost.grid(row=3, column=2, padx=4, sticky="w")

        ttk.Button(form, text="Ajouter ligne", command=self.add_purchase_line).grid(row=3, column=3, padx=5)

        self.purchase_tree = ttk.Treeview(tab, columns=("id", "product", "qty", "cost", "total"), show="headings", height=10)
        for c, t in [("id", "ID"), ("product", "Produit"), ("qty", "Qte"), ("cost", "Cout"), ("total", "Total")]:
            self.purchase_tree.heading(c, text=t)
        self.purchase_tree.pack(fill="both", expand=True, pady=8)

        buttons = ttk.Frame(tab)
        buttons.pack(fill="x")
        ttk.Button(buttons, text="Retirer ligne", command=self.remove_purchase_line).pack(side="left", padx=3)
        ttk.Button(buttons, text="Vider", command=self.clear_purchase_cart).pack(side="left", padx=3)
        ttk.Button(buttons, text="Valider approvisionnement", command=self.finalize_purchase).pack(side="left", padx=6)

    def reload_purchase_dropdowns(self):
        suppliers = self.service.list_suppliers()
        self.suppliers_by_name = {s["name"]: s for s in suppliers}
        if hasattr(self, "purchase_supplier_combo"):
            self.purchase_supplier_combo["values"] = [s["name"] for s in suppliers]

        products = self.service.list_products("")
        self.products_by_purchase_label = {f"{p['name']} ({p['barcode']})": p for p in products}
        if hasattr(self, "purchase_product_combo"):
            self.purchase_product_combo["values"] = list(self.products_by_purchase_label.keys())

    def add_purchase_line(self):
        label = self.purchase_product_combo.get()
        product = self.products_by_purchase_label.get(label)
        if not product:
            return
        try:
            qty = int(self.purchase_qty.get())
            cost = float(self.purchase_cost.get())
            for line in self.purchase_cart:
                if line["product_id"] == product["id"]:
                    line["qty"] += qty
                    line["unit_cost"] = cost
                    self.render_purchase_cart()
                    return
            self.purchase_cart.append({"product_id": product["id"], "name": product["name"], "qty": qty, "unit_cost": cost})
            self.render_purchase_cart()
        except Exception:
            messagebox.showwarning("Approvisionnement", "Quantite/Cout invalides.")

    def render_purchase_cart(self):
        self.purchase_tree.delete(*self.purchase_tree.get_children())
        for line in self.purchase_cart:
            total = line["qty"] * line["unit_cost"]
            self.purchase_tree.insert("", "end", values=(line["product_id"], line["name"], line["qty"], line["unit_cost"], total))
        self._tree_apply_zebra(self.purchase_tree)

    def remove_purchase_line(self):
        sel = self.purchase_tree.selection()
        if not sel:
            return
        pid = int(self.purchase_tree.item(sel[0], "values")[0])
        self.purchase_cart = [x for x in self.purchase_cart if x["product_id"] != pid]
        self.render_purchase_cart()

    def clear_purchase_cart(self):
        self.purchase_cart = []
        self.render_purchase_cart()

    def finalize_purchase(self):
        if not self._require_permission("purchase:manage"):
            return
        supplier_name = self.purchase_supplier_combo.get()
        supplier = self.suppliers_by_name.get(supplier_name)
        if not supplier:
            messagebox.showwarning("Approvisionnement", "Selectionnez un fournisseur.")
            return
        try:
            po = self.service.create_purchase(
                user_id=self.user["id"],
                supplier_id=supplier["id"],
                delivery_date=datetime.strptime(self.purchase_date.get().strip(), "%Y-%m-%d").date(),
                items=self.purchase_cart,
            )
            self.service.audit(self.user["id"], "Approvisionnement", po)
            self.clear_purchase_cart()
            self.refresh_all_data()
            messagebox.showinfo("Approvisionnement", f"Approvisionnement enregistre: {po}")
        except Exception as exc:
            messagebox.showerror("Approvisionnement", str(exc))

    def build_suppliers_tab(self, tab):
        form = ttk.LabelFrame(tab, text="Fournisseur", padding=8)
        form.pack(fill="x")
        self.s_name = self._entry(form, "Nom", 0, 0)
        self.s_phone = self._entry(form, "Telephone", 0, 2)
        self.s_email = self._entry(form, "Email", 2, 0)
        self.s_address = self._entry(form, "Adresse", 2, 2)
        self.s_products = self._entry(form, "Produits fournis", 4, 0)

        row = ttk.Frame(form)
        row.grid(row=6, column=0, columnspan=4, sticky="w", pady=6)
        ttk.Button(row, text="Ajouter", command=self.add_supplier).pack(side="left", padx=3)
        ttk.Button(row, text="Supprimer", command=self.delete_supplier).pack(side="left", padx=3)

        search_row = ttk.Frame(tab)
        search_row.pack(fill="x", pady=(6, 2))
        ttk.Label(search_row, text="Recherche fournisseur").pack(side="left")
        self.suppliers_search = ttk.Entry(search_row, width=36)
        self.suppliers_search.pack(side="left", padx=6)
        self.suppliers_search.bind("<KeyRelease>", lambda _e: self.refresh_suppliers_tree())

        self.suppliers_tree = ttk.Treeview(tab, columns=("id", "name", "phone", "email", "address", "products"), show="headings", height=13)
        for c, t in [
            ("id", "ID"),
            ("name", "Nom"),
            ("phone", "Telephone"),
            ("email", "Email"),
            ("address", "Adresse"),
            ("products", "Produits"),
        ]:
            self.suppliers_tree.heading(c, text=t)
        self.suppliers_tree.pack(fill="both", expand=True, pady=8)

    def add_supplier(self):
        if not self._require_permission("suppliers:manage"):
            return
        try:
            self.service.add_supplier(
                self.s_name.get().strip(),
                self.s_phone.get().strip(),
                self.s_email.get().strip(),
                self.s_address.get().strip(),
                self.s_products.get().strip(),
            )
            self.refresh_suppliers_tree()
            self.reload_purchase_dropdowns()
        except Exception as exc:
            messagebox.showerror("Fournisseurs", str(exc))

    def delete_supplier(self):
        if not self._require_permission("suppliers:manage"):
            return
        sel = self.suppliers_tree.selection()
        if not sel:
            return
        supplier_id = int(self.suppliers_tree.item(sel[0], "values")[0])
        try:
            self.service.delete_supplier(supplier_id)
            self.refresh_suppliers_tree()
            self.reload_purchase_dropdowns()
        except Exception as exc:
            messagebox.showerror("Fournisseurs", str(exc))

    def refresh_suppliers_tree(self):
        self.suppliers_tree.delete(*self.suppliers_tree.get_children())
        term = self.suppliers_search.get().strip().lower() if hasattr(self, "suppliers_search") else ""
        for row in self.service.list_suppliers():
            if term:
                blob = f"{row['name']} {row['phone']} {row['email']} {row['supplied_products']}".lower()
                if term not in blob:
                    continue
            self.suppliers_tree.insert(
                "",
                "end",
                values=(row["id"], row["name"], row["phone"], row["email"], row["address"], row["supplied_products"]),
            )
        self._tree_apply_zebra(self.suppliers_tree)

    def build_clients_tab(self, tab):
        form = ttk.LabelFrame(tab, text="Client", padding=8)
        form.pack(fill="x")
        self.c_name = self._entry(form, "Nom", 0, 0)
        self.c_phone = self._entry(form, "Telephone", 0, 2)
        self.c_email = self._entry(form, "Email", 2, 0)
        self.c_address = self._entry(form, "Adresse", 2, 2)

        row = ttk.Frame(form)
        row.grid(row=4, column=0, columnspan=4, sticky="w", pady=6)
        ttk.Button(row, text="Ajouter", command=self.add_client).pack(side="left", padx=3)
        ttk.Button(row, text="Supprimer", command=self.delete_client).pack(side="left", padx=3)

        search_row = ttk.Frame(tab)
        search_row.pack(fill="x", pady=(6, 2))
        ttk.Label(search_row, text="Recherche client").pack(side="left")
        self.clients_search = ttk.Entry(search_row, width=36)
        self.clients_search.pack(side="left", padx=6)
        self.clients_search.bind("<KeyRelease>", lambda _e: self.refresh_clients_tree())

        self.clients_tree = ttk.Treeview(tab, columns=("id", "name", "phone", "email", "address"), show="headings", height=15)
        for c, t in [("id", "ID"), ("name", "Nom"), ("phone", "Telephone"), ("email", "Email"), ("address", "Adresse")]:
            self.clients_tree.heading(c, text=t)
        self.clients_tree.pack(fill="both", expand=True, pady=8)

    def add_client(self):
        if not self._require_permission("clients:manage"):
            return
        try:
            self.service.add_client(
                self.c_name.get().strip(),
                self.c_phone.get().strip(),
                self.c_email.get().strip(),
                self.c_address.get().strip(),
            )
            self.refresh_clients_tree()
            self.reload_pos_dropdowns()
        except Exception as exc:
            messagebox.showerror("Clients", str(exc))

    def delete_client(self):
        if not self._require_permission("clients:manage"):
            return
        sel = self.clients_tree.selection()
        if not sel:
            return
        client_id = int(self.clients_tree.item(sel[0], "values")[0])
        try:
            self.service.delete_client(client_id)
            self.refresh_clients_tree()
            self.reload_pos_dropdowns()
        except Exception as exc:
            messagebox.showerror("Clients", str(exc))

    def refresh_clients_tree(self):
        self.clients_tree.delete(*self.clients_tree.get_children())
        term = self.clients_search.get().strip().lower() if hasattr(self, "clients_search") else ""
        for row in self.service.list_clients():
            if term:
                blob = f"{row['full_name']} {row['phone']} {row['email']} {row['address']}".lower()
                if term not in blob:
                    continue
            self.clients_tree.insert("", "end", values=(row["id"], row["full_name"], row["phone"], row["email"], row["address"]))
        self._tree_apply_zebra(self.clients_tree)

    def build_reports_tab(self, tab):
        box = ttk.Frame(tab)
        box.pack(fill="x")

        ttk.Label(box, text="Du").pack(side="left")
        self.rep_from = ttk.Entry(box, width=14)
        self.rep_from.insert(0, (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
        self.rep_from.pack(side="left", padx=4)

        ttk.Label(box, text="Au").pack(side="left")
        self.rep_to = ttk.Entry(box, width=14)
        self.rep_to.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.rep_to.pack(side="left", padx=4)

        ttk.Button(box, text="Generer", command=self.refresh_reports).pack(side="left", padx=5)
        ttk.Button(box, text="Export CSV (Excel)", command=self.export_sales_csv).pack(side="left", padx=5)
        ttk.Button(box, text="Export Excel avance", command=self.export_reports_workbook).pack(side="left", padx=5)
        ttk.Button(box, text="Export PDF", command=self.export_sales_pdf).pack(side="left", padx=5)

        self.lbl_report_revenue = ttk.Label(tab, text=f"CA: {self._money(0)}", font=("Segoe UI", 11, "bold"))
        self.lbl_report_revenue.pack(anchor="w", pady=6)
        self.lbl_report_profit = ttk.Label(tab, text=f"Benefice brut: {self._money(0)}", font=("Segoe UI", 11, "bold"))
        self.lbl_report_profit.pack(anchor="w", pady=2)

        self.report_tree = ttk.Treeview(tab, columns=("invoice", "total", "payment", "date"), show="headings", height=15)
        for c, t in [("invoice", "Facture"), ("total", "Total"), ("payment", "Paiement"), ("date", "Date")]:
            self.report_tree.heading(c, text=t)
        self.report_tree.pack(fill="both", expand=True)

    def refresh_reports(self):
        from_date = self.rep_from.get().strip()
        to_date = self.rep_to.get().strip()
        rows = self.service.sales_report(from_date, to_date)
        self.current_report_rows = rows

        self.report_tree.delete(*self.report_tree.get_children())
        for row in rows:
            self.report_tree.insert("", "end", values=(row["invoice_number"], row["total_amount"], row["payment_mode"], row["created_at"]))
        self._tree_apply_zebra(self.report_tree)

        summary = self.service.finance_summary(from_date, to_date)
        self.lbl_report_revenue.config(text=f"CA: {self._money(float(summary['revenue']))}")
        self.lbl_report_profit.config(text=f"Benefice brut: {self._money(float(summary['gross_profit']))}")

    def export_sales_csv(self):
        if not self._require_permission("reports:view"):
            return
        if not getattr(self, "current_report_rows", None):
            return
        out_dir = EXPORTS_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / f"rapport_ventes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Facture", "Total", "Paiement", "Date"])
            for row in self.current_report_rows:
                writer.writerow([row["invoice_number"], row["total_amount"], row["payment_mode"], row["created_at"]])
        messagebox.showinfo("Rapports", f"Export CSV cree: {path}")

    def export_sales_pdf(self):
        if not self._require_permission("reports:view"):
            return
        rows = getattr(self, "current_report_rows", [])
        if not rows:
            return
        out_dir = EXPORTS_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / f"rapport_ventes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        c = canvas.Canvas(str(path), pagesize=A4)
        width, height = A4
        y = height - 40
        c.setFillColor(colors.HexColor(PALETTE["primary"]))
        c.rect(0, height - 62, width, 62, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y + 4, "Rapport des ventes")
        c.setFont("Helvetica", 9)
        c.drawString(40, y - 10, f"Periode: {self.rep_from.get()} -> {self.rep_to.get()}")
        y -= 20

        c.setFillColor(colors.HexColor("#f6f9ff"))
        c.roundRect(34, y - 8, width - 68, 30 + (len(rows) * 12), 6, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#e8f0ff"))
        c.rect(36, y - 5, width - 72, 18, fill=1, stroke=0)

        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.HexColor("#183153"))
        c.drawString(40, y, "Facture")
        c.drawString(220, y, "Total")
        c.drawString(290, y, "Paiement")
        c.drawString(420, y, "Date")
        y -= 14

        c.setFont("Helvetica", 9)
        for idx, row in enumerate(rows):
            if y < 60:
                c.showPage()
                y = height - 40
            if idx % 2 == 0:
                c.setFillColor(colors.HexColor("#f0f6ff"))
                c.rect(36, y - 3, width - 72, 12, fill=1, stroke=0)
            c.setFillColor(colors.HexColor("#1e2f45"))
            c.drawString(40, y, str(row["invoice_number"]))
            c.drawString(220, y, self._money(float(row['total_amount'])))
            pay_mode = str(row["payment_mode"])
            pay_color = colors.HexColor("#34a853") if "Mobile" in pay_mode else (colors.HexColor("#1a73e8") if "Carte" in pay_mode else colors.HexColor("#5f6368"))
            c.setFillColor(pay_color)
            c.roundRect(288, y - 2, 90, 10, 4, fill=1, stroke=0)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 7)
            c.drawString(292, y, pay_mode[:18])
            c.setFillColor(colors.HexColor("#1e2f45"))
            c.setFont("Helvetica", 9)
            c.drawString(420, y, str(row["created_at"]))
            y -= 12

        y -= 12
        total_revenue = sum(float(r["total_amount"]) for r in rows)
        c.setFillColor(colors.HexColor(PALETTE["primary"]))
        c.roundRect(360, y - 12, 162, 24, 6, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(370, y, f"Total periode: {self._money(total_revenue)}")
        c.save()
        messagebox.showinfo("Rapports", f"Export PDF cree: {path}")

    def export_reports_workbook(self):
        if not self._require_permission("reports:view"):
            return
        from_date = self.rep_from.get().strip()
        to_date = self.rep_to.get().strip()
        sales_rows = self.service.sales_report(from_date, to_date)
        stock_rows = self.service.stock_alerts()
        summary = self.service.finance_summary(from_date, to_date)

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Ventes"
        ws1.append(["Facture", "Total", "Paiement", "Date"])
        for r in sales_rows:
            ws1.append([r["invoice_number"], float(r["total_amount"]), r["payment_mode"], str(r["created_at"])])

        ws2 = wb.create_sheet("Stock")
        ws2.append(["Produit", "Code", "Stock", "Min"])
        for r in stock_rows:
            ws2.append([r["name"], r["barcode"], int(r["stock_qty"]), int(r["min_stock"])])

        ws3 = wb.create_sheet("Financier")
        ws3.append(["Periode", f"{from_date} -> {to_date}"])
        ws3.append(["Chiffre d'affaires", float(summary["revenue"])])
        ws3.append(["Benefice brut", float(summary["gross_profit"])])

        out_dir = EXPORTS_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / f"rapport_complet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(path)
        messagebox.showinfo("Rapports", f"Export Excel cree: {path}")

    def build_users_tab(self, tab):
        if "users:manage" not in self.permissions:
            ttk.Label(tab, text="Acces reserve a l'administrateur.").pack(anchor="w")
            return

        form = ttk.LabelFrame(tab, text="Nouvel utilisateur", padding=8)
        form.pack(fill="x")

        self.u_fullname = self._entry(form, "Nom complet", 0, 0)
        self.u_username = self._entry(form, "Identifiant", 0, 2)
        self.u_password = self._entry(form, "Mot de passe", 2, 0, show="*")

        ttk.Label(form, text="Role").grid(row=2, column=2, sticky="w")
        self.u_role = ttk.Combobox(form, state="readonly", width=22)
        self.u_role.grid(row=3, column=2)

        ttk.Button(form, text="Ajouter utilisateur", command=self.add_user).grid(row=4, column=0, pady=6, sticky="w")
        ttk.Button(form, text="Activer/Desactiver", command=self.toggle_user).grid(row=4, column=1, pady=6, sticky="w")
        ttk.Button(form, text="Supprimer utilisateur", command=self.delete_user, style="Danger.TButton").grid(row=4, column=2, pady=6, sticky="w")

        search_row = ttk.Frame(tab)
        search_row.pack(fill="x", pady=(6, 2))
        ttk.Label(search_row, text="Recherche utilisateur").pack(side="left")
        self.users_search = ttk.Entry(search_row, width=36)
        self.users_search.pack(side="left", padx=6)
        self.users_search.bind("<KeyRelease>", lambda _e: self.refresh_users_tree())

        self.users_tree = ttk.Treeview(tab, columns=("id", "name", "username", "role", "active"), show="headings", height=14)
        for c, t in [("id", "ID"), ("name", "Nom"), ("username", "Identifiant"), ("role", "Role"), ("active", "Actif")]:
            self.users_tree.heading(c, text=t)
        self.users_tree.pack(fill="both", expand=True, pady=8)

    def refresh_users_tree(self):
        if not hasattr(self, "users_tree"):
            return
        roles = self.service.list_roles()
        self.role_by_name = {r["name"]: r["id"] for r in roles}
        self.u_role["values"] = [r["name"] for r in roles]
        if roles and not self.u_role.get():
            self.u_role.set(roles[0]["name"])

        self.users_tree.delete(*self.users_tree.get_children())
        term = self.users_search.get().strip().lower() if hasattr(self, "users_search") else ""
        for row in self.service.list_users():
            if term:
                blob = f"{row['full_name']} {row['username']} {row['role_name']}".lower()
                if term not in blob:
                    continue
            self.users_tree.insert(
                "",
                "end",
                values=(row["id"], row["full_name"], row["username"], row["role_name"], "Oui" if row["is_active"] else "Non"),
            )
        self._tree_apply_zebra(self.users_tree)

    def add_user(self):
        if not self._require_permission("users:manage"):
            return
        try:
            self.service.add_user(
                self.u_fullname.get().strip(),
                self.u_username.get().strip(),
                self.u_password.get().strip(),
                self.role_by_name[self.u_role.get()],
            )
            self.refresh_users_tree()
            messagebox.showinfo("Utilisateurs", "Utilisateur ajoute.")
        except Exception as exc:
            messagebox.showerror("Utilisateurs", str(exc))

    def toggle_user(self):
        if not self._require_permission("users:manage"):
            return
        sel = self.users_tree.selection()
        if not sel:
            return
        vals = self.users_tree.item(sel[0], "values")
        uid = int(vals[0])
        current_active = vals[4] == "Oui"
        try:
            self.service.toggle_user(uid, not current_active)
            self.refresh_users_tree()
        except Exception as exc:
            messagebox.showerror("Utilisateurs", str(exc))

    def delete_user(self):
        if not self._require_permission("users:manage"):
            return
        sel = self.users_tree.selection()
        if not sel:
            return
        vals = self.users_tree.item(sel[0], "values")
        uid = int(vals[0])
        username = vals[2]
        if uid == self.user["id"]:
            messagebox.showwarning("Utilisateurs", "Vous ne pouvez pas supprimer votre session en cours.")
            return
        if not messagebox.askyesno("Utilisateurs", f"Supprimer l'utilisateur {username} ?"):
            return
        try:
            self.service.delete_user(uid)
            self.service.audit(self.user["id"], "Utilisateur supprime", f"ID {uid}")
            self.refresh_users_tree()
            messagebox.showinfo("Utilisateurs", "Utilisateur supprime.")
        except Exception as exc:
            messagebox.showerror("Utilisateurs", str(exc))

    def build_settings_tab(self, tab):
        ttk.Label(tab, text="Design et branding", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=4)
        theme_row = ttk.Frame(tab)
        theme_row.pack(fill="x", pady=4)
        ttk.Label(theme_row, text="Theme").pack(side="left")
        self.theme_combo = ttk.Combobox(theme_row, state="readonly", width=22)
        self.theme_combo["values"] = list(THEME_PRESETS.keys())
        self.theme_combo.set(self.current_theme if self.current_theme in THEME_PRESETS else "Clair Pro")
        self.theme_combo.pack(side="left", padx=6)
        ttk.Button(theme_row, text="Appliquer theme", command=self.apply_ui_theme).pack(side="left", padx=6)

        brand_row = ttk.Frame(tab)
        brand_row.pack(fill="x", pady=4)
        ttk.Label(brand_row, text="Nom magasin").pack(side="left")
        self.brand_entry = ttk.Entry(brand_row, width=30)
        self.brand_entry.insert(0, self.brand_name)
        self.brand_entry.pack(side="left", padx=6)
        ttk.Button(brand_row, text="Enregistrer branding", command=self.save_branding).pack(side="left", padx=6)

        currency_row = ttk.Frame(tab)
        currency_row.pack(fill="x", pady=4)
        ttk.Label(currency_row, text="Devise").pack(side="left")
        self.currency_combo = ttk.Combobox(currency_row, state="readonly", width=36)
        self.currency_combo["values"] = [f"{code} - {name}" for code, name in CURRENCY_CHOICES]
        self.currency_combo.set(self._currency_display(self.currency_code))
        self.currency_combo.pack(side="left", padx=6)
        ttk.Button(currency_row, text="Appliquer devise", command=self.save_currency).pack(side="left", padx=6)

        ttk.Separator(tab).pack(fill="x", pady=10)
        ttk.Label(tab, text="Contexte magasin et caisse", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=4)
        ctx = ttk.Frame(tab)
        ctx.pack(fill="x", pady=6)

        ttk.Label(ctx, text="Magasin").pack(side="left")
        self.store_combo = ttk.Combobox(ctx, state="readonly", width=28)
        self.store_combo.pack(side="left", padx=6)
        self.store_combo.bind("<<ComboboxSelected>>", lambda _e: self.reload_registers_for_store())

        ttk.Label(ctx, text="Caisse").pack(side="left", padx=(10, 0))
        self.register_combo = ttk.Combobox(ctx, state="readonly", width=22)
        self.register_combo.pack(side="left", padx=6)
        ttk.Button(ctx, text="Appliquer contexte", command=self.apply_runtime_context).pack(side="left", padx=8)

        ttk.Separator(tab).pack(fill="x", pady=10)
        ttk.Label(tab, text="Sauvegarde et restauration", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=4)
        btns = ttk.Frame(tab)
        btns.pack(anchor="w", pady=8)
        ttk.Button(btns, text="Sauvegarde manuelle", command=self.backup_data).pack(side="left", padx=4)
        ttk.Button(btns, text="Restauration complete", command=self.restore_data, style="Warning.TButton").pack(side="left", padx=4)

        ttk.Separator(tab).pack(fill="x", pady=10)
        ttk.Label(tab, text="Securite compte", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=4)
        sec = ttk.Frame(tab)
        sec.pack(fill="x", pady=4)
        ttk.Label(sec, text="Ancien mot de passe").grid(row=0, column=0, sticky="w")
        self.old_pwd = ttk.Entry(sec, width=24, show="*")
        self.old_pwd.grid(row=1, column=0, padx=4, pady=2)
        ttk.Label(sec, text="Nouveau mot de passe").grid(row=0, column=1, sticky="w")
        self.new_pwd = ttk.Entry(sec, width=24, show="*")
        self.new_pwd.grid(row=1, column=1, padx=4, pady=2)
        ttk.Button(sec, text="Changer mot de passe", command=self.change_my_password).grid(row=1, column=2, padx=8)

        ttk.Label(tab, text="Politique: min 8 caracteres, majuscule, minuscule, chiffre.", foreground="#555").pack(anchor="w", pady=(6, 2))

        self.load_settings_context()

    def apply_ui_theme(self):
        if not self._require_permission("settings:manage"):
            return
        theme_name = self.theme_combo.get()
        if theme_name not in THEME_PRESETS:
            return
        self.service.set_setting("ui.theme", theme_name)
        self.current_theme = theme_name
        self.app.apply_theme(theme_name)
        messagebox.showinfo("Design", f"Theme applique: {theme_name}")

    def save_branding(self):
        if not self._require_permission("settings:manage"):
            return
        name = self.brand_entry.get().strip() or "Magasin POS"
        self.service.set_setting("brand.name", name)
        self.brand_name = name
        self.top_title.config(text=f"{self.brand_name}  •  Point de Vente")
        messagebox.showinfo("Design", "Branding enregistre.")

    def save_currency(self):
        if not self._require_permission("settings:manage"):
            return
        selected = self.currency_combo.get().strip()
        code = selected.split(" - ", 1)[0].upper() if selected else "EUR"
        valid_codes = {c for c, _ in CURRENCY_CHOICES}
        if code not in valid_codes:
            messagebox.showwarning("Parametres", "Devise invalide.")
            return
        self.currency_code = code
        self.service.set_setting("app.currency", code)
        if hasattr(self, "lbl_revenue"):
            self.refresh_dashboard()
        if hasattr(self, "lbl_cart_total"):
            self.render_cart()
        if hasattr(self, "lbl_report_revenue"):
            self.refresh_reports()
        messagebox.showinfo("Parametres", f"Devise appliquee: {self._currency_display(code)}")

    def backup_data(self):
        if not self._require_permission("settings:manage"):
            return
        tables = [
            "roles",
            "users",
            "categories",
            "suppliers",
            "clients",
            "products",
            "sales",
            "sale_items",
            "purchase_orders",
            "purchase_items",
            "stock_movements",
            "returns",
            "audit_logs",
        ]
        payload = {}
        for table in tables:
            payload[table] = fetch_all(f"SELECT * FROM {table}")

        out_dir = BACKUPS_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        file_path = out_dir / f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with file_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, default=str, ensure_ascii=False, indent=2)
        messagebox.showinfo("Sauvegarde", f"Sauvegarde creee: {file_path}")

    def restore_data(self):
        if not self._require_permission("settings:manage"):
            return
        file_path = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if not file_path:
            return
        if not messagebox.askyesno("Restauration", "Cette action remplacera les donnees actuelles. Continuer ?"):
            return
        try:
            self.service.restore_backup_json(file_path)
            self.service.ensure_default_admin()
            self.refresh_all_data()
            messagebox.showinfo("Restauration", "Restauration terminee avec succes.")
        except Exception as exc:
            messagebox.showerror("Restauration", str(exc))

    def load_settings_context(self):
        if not hasattr(self, "store_combo"):
            return
        stores = self.service.list_stores()
        self.stores_by_name = {s["name"]: s for s in stores}
        self.store_combo["values"] = list(self.stores_by_name.keys())

        runtime = self.service.get_runtime_context()
        target_store = next((s["name"] for s in stores if s["id"] == runtime["store_id"]), stores[0]["name"] if stores else "")
        if target_store:
            self.store_combo.set(target_store)
            self.reload_registers_for_store(preferred_id=runtime["register_id"])

    def reload_registers_for_store(self, preferred_id: int | None = None):
        store_name = self.store_combo.get()
        store = self.stores_by_name.get(store_name)
        if not store:
            return
        regs = self.service.list_registers(store["id"])
        self.registers_by_name = {r["name"]: r for r in regs}
        self.register_combo["values"] = list(self.registers_by_name.keys())
        if preferred_id:
            pref = next((r["name"] for r in regs if r["id"] == preferred_id), "")
            if pref:
                self.register_combo.set(pref)
        if not self.register_combo.get() and regs:
            self.register_combo.set(regs[0]["name"])

    def apply_runtime_context(self):
        if not self._require_permission("settings:manage"):
            return
        store = self.stores_by_name.get(self.store_combo.get())
        register = self.registers_by_name.get(self.register_combo.get())
        if not store or not register:
            messagebox.showwarning("Parametres", "Selection magasin/caisse invalide.")
            return
        self.service.set_runtime_context(store["id"], register["id"])
        self.runtime_context = self.service.get_runtime_context()
        self.lbl_store_context.config(text=f"Magasin {self.runtime_context['store_id']} / Caisse {self.runtime_context['register_id']}")
        messagebox.showinfo("Parametres", "Contexte applique.")

    def change_my_password(self):
        old = self.old_pwd.get().strip()
        new = self.new_pwd.get().strip()
        if not old or not new:
            return
        try:
            self.service.change_password(self.user["id"], old, new)
            self.old_pwd.delete(0, "end")
            self.new_pwd.delete(0, "end")
            messagebox.showinfo("Securite", "Mot de passe mis a jour.")
        except Exception as exc:
            messagebox.showerror("Securite", str(exc))

    def _currency_display(self, code: str) -> str:
        for c, name in CURRENCY_CHOICES:
            if c == code:
                return f"{c} - {name}"
        return "EUR - Euro"

    def _money(self, value: float) -> str:
        return f"{float(value):.2f} {self.currency_code}"

    def _metric_card(self, parent, title: str, value: str, color: str, column: int):
        card = tk.Frame(parent, bg=color, padx=14, pady=12)
        card.grid(row=0, column=column, padx=8, pady=4, sticky="nsew")
        parent.grid_columnconfigure(column, weight=1)
        if hasattr(self, "dashboard_cards"):
            self.dashboard_cards.append(card)

        tk.Label(card, text=title, bg=color, fg="#f6fbff", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        value_label = tk.Label(card, text=value, bg=color, fg="#ffffff", font=("Segoe UI", 20, "bold"))
        value_label.pack(anchor="w", pady=(6, 0))
        spark = tk.Canvas(card, width=168, height=38, bg=color, highlightthickness=0)
        spark.pack(anchor="w", pady=(6, 0))
        return value_label, spark

    def _arrange_dashboard_cards(self, width: int):
        if not hasattr(self, "dashboard_cards") or not self.dashboard_cards:
            return

        if width >= 1400:
            cols = 5
        elif width >= 1240:
            cols = 4
        elif width >= 1080:
            cols = 3
        elif width >= 900:
            cols = 2
        else:
            cols = 1

        container = self.dashboard_cards_container
        for i in range(6):
            container.grid_columnconfigure(i, weight=0)

        for idx, card in enumerate(self.dashboard_cards):
            r = idx // cols
            c = idx % cols
            container.grid_columnconfigure(c, weight=1)
            card.grid_configure(row=r, column=c, padx=8, pady=6, sticky="nsew")

    def _apply_pos_density(self, compact: bool):
        if not hasattr(self, "pos_barcode"):
            return

        barcode_w = 24 if compact else 30
        combo_w = 22 if compact else 30
        client_w = 18 if compact else 24
        amount_w = 10 if compact else 14
        pay_w = 18 if compact else 24

        self.pos_barcode.configure(width=barcode_w)
        self.pos_product_combo.configure(width=combo_w)
        self.pos_client_combo.configure(width=client_w)
        self.pos_discount.configure(width=amount_w)
        self.pos_vat.configure(width=amount_w)
        self.pos_payment.configure(width=pay_w)

        if hasattr(self, "cart_tree"):
            self.cart_tree.configure(height=14 if compact else 18)

        if hasattr(self, "pos_head"):
            self.pos_head.configure(padding=6 if compact else 8)
        if hasattr(self, "pos_summary"):
            self.pos_summary.configure(padding=8 if compact else 10)
        if hasattr(self, "pos_pay"):
            self.pos_pay.configure(padding=8 if compact else 10)

    def _animate_numeric_label(self, label: tk.Label, target: float, suffix: str, decimals: int = 0):
        steps = 10
        current = getattr(label, "_last_value", 0.0)
        delta = (target - current) / steps if steps else 0

        def fmt(v: float) -> str:
            if decimals <= 0:
                return f"{int(round(v))}{suffix}"
            return f"{v:.{decimals}f}{suffix}"

        def run(i: int, value: float):
            if i >= steps:
                label.config(text=fmt(target))
                label._last_value = target
                return
            nxt = value + delta
            label.config(text=fmt(nxt))
            self.after(22, lambda: run(i + 1, nxt))

        run(0, current)

    def _tree_apply_zebra(self, tree: ttk.Treeview):
        if PALETTE["bg"] == "#151c2b":
            even, odd = "#233049", "#1f2738"
        else:
            even, odd = "#f8fbff", "#ffffff"
        tree.tag_configure("row_even", background=even)
        tree.tag_configure("row_odd", background=odd)
        for i, item in enumerate(tree.get_children()):
            tags = list(tree.item(item, "tags"))
            tags = [t for t in tags if t not in {"row_even", "row_odd"}]
            tags.append("row_even" if i % 2 == 0 else "row_odd")
            tree.item(item, tags=tuple(tags))

    def _apply_compact_mode(self, compact: bool):
        if compact == self._compact_mode:
            return
        self._compact_mode = compact

        style = ttk.Style(self)
        if compact:
            style.configure("TButton", font=("Segoe UI", 8, "bold"), padding=(7, 4))
            style.configure("Nav.TButton", font=("Segoe UI", 8, "bold"), padding=(8, 7))
            style.configure("NavActive.TButton", font=("Segoe UI", 8, "bold"), padding=(8, 7))
            style.configure("TLabel", font=("Segoe UI", 9))
            style.configure("Treeview", rowheight=21)
            self.top_bar.configure(padding=(10, 5))
        else:
            style.configure("TButton", font=("Segoe UI", 9, "bold"), padding=(10, 6))
            style.configure("Nav.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 10))
            style.configure("NavActive.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 10))
            style.configure("TLabel", font=("Segoe UI", 10))
            style.configure("Treeview", rowheight=25)
            self.top_bar.configure(padding=(14, 8))

    def _apply_sidebar_density(self, compact: bool):
        style = ttk.Style(self)
        if compact:
            style.configure("Nav.TButton", font=("Segoe UI", 8, "bold"), padding=(8, 5))
            style.configure("NavActive.TButton", font=("Segoe UI", 8, "bold"), padding=(8, 5))
            style.configure("Nav.TLabel", font=("Segoe UI", 8, "bold"))
            self.sidebar_header.configure(padx=10, pady=10)
            for lbl in getattr(self, "nav_group_labels", []):
                lbl.pack_configure(padx=10, pady=(6, 2))
        else:
            style.configure("Nav.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 10))
            style.configure("NavActive.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 10))
            style.configure("Nav.TLabel", font=("Segoe UI", 9, "bold"))
            self.sidebar_header.configure(padx=12, pady=16)
            for lbl in getattr(self, "nav_group_labels", []):
                lbl.pack_configure(padx=12, pady=(10, 4))

    def _apply_pos_layout(self, vertical: bool):
        if not hasattr(self, "pos_left") or not hasattr(self, "pos_right"):
            return
        if vertical == self._pos_vertical_mode:
            return
        self._pos_vertical_mode = vertical

        self.pos_left.pack_forget()
        self.pos_right.pack_forget()
        if vertical:
            self.pos_left.pack(side="top", fill="both", expand=True, padx=0, pady=(0, 8))
            self.pos_right.pack(side="top", fill="x")
        else:
            self.pos_left.pack(side="left", fill="both", expand=True, padx=(0, 8), pady=0)
            self.pos_right.pack(side="left", fill="y")

    def _apply_table_compaction(self, compact: bool):
        specs = {
            "products_tree": {
                "id": (50, 34),
                "nom": (200, 120),
                "barcode": (140, 95),
                "categorie": (120, 90),
                "achat": (90, 70),
                "vente": (90, 70),
                "marque": (120, 90),
                "stock": (70, 56),
                "min": (70, 56),
            },
            "cart_tree": {
                "id": (55, 40),
                "name": (260, 140),
                "qty": (70, 55),
                "price": (95, 70),
                "total": (100, 76),
            },
            "report_tree": {
                "invoice": (160, 120),
                "total": (95, 75),
                "payment": (130, 98),
                "date": (180, 130),
            },
            "invoices_tree": {
                "invoice": (160, 120),
                "total": (90, 70),
                "payment": (120, 95),
                "date": (170, 125),
            },
            "stock_move_tree": {
                "date": (160, 120),
                "product": (180, 130),
                "type": (120, 90),
                "qty": (90, 70),
                "note": (260, 130),
            },
        }

        for tree_name, cols in specs.items():
            tree = getattr(self, tree_name, None)
            if not tree:
                continue
            for col_name, (normal_w, compact_w) in cols.items():
                try:
                    tree.column(col_name, width=compact_w if compact else normal_w, minwidth=40, stretch=True)
                except tk.TclError:
                    continue

    def _entry(self, parent, label: str, row: int, col: int, show: str | None = None):
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", pady=3)
        e = ttk.Entry(parent, width=28, show=show if show else "")
        e.grid(row=row + 1, column=col, padx=4, sticky="w")
        return e

    def _set_entry(self, entry: ttk.Entry, value):
        entry.delete(0, "end")
        entry.insert(0, str(value) if value is not None else "")


if __name__ == "__main__":
    app = StoreApp()
    app.mainloop()
